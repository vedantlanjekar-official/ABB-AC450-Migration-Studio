"""
report_generator.py — Build Comparison_Report.xlsx for Excel Comparison & Validation.

Sheets:
  1. Summary — totals for matched / unmatched
  2. Unmatched Records — tags unique to either source workbook
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font

from backend.excel.design import build_db_excel_design
from backend.excel_compare.comparator import ComparisonResult


class ComparisonReportGenerator:
    """Writes a professional two-sheet comparison workbook."""

    def generate(self, result: ComparisonResult, output_path: Path | str) -> Path:
        path = Path(output_path)
        if path.parent:
            path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        design = build_db_excel_design()
        title_font = Font(name="Calibri", size=14, bold=True, color="1E293B")
        subtitle_font = Font(
            name="Calibri",
            size=9,
            italic=True,
            color="475569",
        )
        label_font = Font(name="Calibri", size=10, bold=True, color="000000")

        # --- Sheet 1: Summary ---
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary.merge_cells("A1:B1")
        ws_summary["A1"] = "Excel Comparison & Validation Report"
        ws_summary["A1"].font = title_font
        ws_summary["A1"].alignment = design.left_align

        ws_summary["A2"] = f"Excel 1 ($(DEVICETAG)): {result.worksheet1_file}"
        ws_summary["A3"] = f"Excel 2 ($(DEVICETAG)): {result.worksheet2_file}"
        ws_summary["A2"].font = subtitle_font
        ws_summary["A3"].font = subtitle_font

        headers = ["Metric", "Value"]
        for col, text in enumerate(headers, start=1):
            cell = ws_summary.cell(row=5, column=col, value=text)
            cell.fill = design.header_fill
            cell.font = design.header_font
            cell.alignment = design.center_align
            cell.border = design.thin_border

        summary_rows = [
            ("Worksheet 1 Records", result.worksheet1_records),
            ("Worksheet 2 Records", result.worksheet2_records),
            ("Matched Records", result.matched_records),
            ("Total Unmatched Records", result.unmatched_records),
            ("Only in Excel 1", len(result.unmatched_in_worksheet1)),
            ("Only in Excel 2", len(result.unmatched_in_worksheet2)),
            ("WS1 Duplicates Ignored", result.worksheet1_duplicates),
            ("WS2 Duplicates Ignored", result.worksheet2_duplicates),
        ]
        for idx, (metric, value) in enumerate(summary_rows):
            row = 6 + idx
            c1 = ws_summary.cell(row=row, column=1, value=metric)
            c2 = ws_summary.cell(row=row, column=2, value=value)
            c1.font = label_font
            c2.font = design.cell_font
            c1.alignment = design.left_align
            c2.alignment = design.center_align
            c1.border = design.thin_border
            c2.border = design.thin_border
            row_fill = design.zebra_fill if idx % 2 == 1 else design.white_fill
            c1.fill = row_fill
            c2.fill = row_fill

        ws_summary.column_dimensions["A"].width = 28
        ws_summary.column_dimensions["B"].width = 14

        # --- Sheet 2: Unmatched Records ---
        ws_unmatched = wb.create_sheet("Unmatched Records")
        ws_unmatched.merge_cells("A1:D1")
        ws_unmatched["A1"] = "Unmatched $(DEVICETAG) Values from Both Excel Files"
        ws_unmatched["A1"].font = title_font

        unmatched_headers = [
            "Sr. No.",
            "$(DEVICETAG)",
            "Source Excel",
            "Missing From",
        ]
        for col, text in enumerate(unmatched_headers, start=1):
            cell = ws_unmatched.cell(row=3, column=col, value=text)
            cell.fill = design.header_fill
            cell.font = design.header_font
            cell.alignment = design.center_align
            cell.border = design.thin_border

        if result.unmatched_items:
            for i, item in enumerate(result.unmatched_items, start=1):
                row = 3 + i
                values = [
                    i,
                    item.tag,
                    item.source_file,
                    item.missing_from_file,
                ]
                for col, val in enumerate(values, start=1):
                    cell = ws_unmatched.cell(row=row, column=col, value=val)
                    cell.font = design.cell_font
                    cell.border = design.thin_border
                    cell.alignment = (
                        design.center_align if col != 2 else design.left_align
                    )
                    cell.fill = (
                        design.zebra_fill if i % 2 == 0 else design.white_fill
                    )
        else:
            values = ["—", "All $(DEVICETAG) values matched", "—", "—"]
            for col, value in enumerate(values, start=1):
                c = ws_unmatched.cell(row=4, column=col, value=value)
                c.font = design.cell_font
                c.border = design.thin_border
                c.fill = design.white_fill

        ws_unmatched.column_dimensions["A"].width = 10
        ws_unmatched.column_dimensions["B"].width = 32
        ws_unmatched.column_dimensions["C"].width = 32
        ws_unmatched.column_dimensions["D"].width = 32

        wb.save(path)
        return path

    @staticmethod
    def build_preview(result: ComparisonResult) -> Dict[str, List[Dict[str, Any]]]:
        """Preview payload for the frontend results view."""
        summary = [
            {"Metric": "Worksheet 1 Records", "Value": result.worksheet1_records},
            {"Metric": "Worksheet 2 Records", "Value": result.worksheet2_records},
            {"Metric": "Matched Records", "Value": result.matched_records},
            {"Metric": "Total Unmatched Records", "Value": result.unmatched_records},
            {
                "Metric": "Only in Excel 1",
                "Value": len(result.unmatched_in_worksheet1),
            },
            {
                "Metric": "Only in Excel 2",
                "Value": len(result.unmatched_in_worksheet2),
            },
        ]
        unmatched = [
            {
                "Sr. No.": i,
                "$(DEVICETAG)": item.tag,
                "Source Excel": item.source_file,
                "Missing From": item.missing_from_file,
            }
            for i, item in enumerate(result.unmatched_items, start=1)
        ]
        return {
            "Summary": summary,
            "Unmatched Records": unmatched,
        }
