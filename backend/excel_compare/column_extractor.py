"""
column_extractor.py — Dynamically locate Excel header columns and extract tag values.

Both workbooks are searched dynamically for the engineering
``$(DEVICETAG)`` column, regardless of worksheet, row, or column position.

Headers and worksheets are detected dynamically — never assume fixed cells or tab order.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple, Any
import logging
import openpyxl


logger = logging.getLogger("excel_compare")

HEADER_SCAN_MAX_ROWS = 120
HEADER_SCAN_MAX_COLS = 50
PREFERRED_SHEET_NAMES = ("I_O_List", "Clubbed_IO", "Unmatched Records", "Summary")
HEADER_ALIASES = {
    "DEVICE TAG": frozenset({"DEVICE TAG", "$(DEVICETAG)"}),
    "NAME": frozenset({"NAME", "$(DEVICETAG)"}),
}


@dataclass
class ExtractionResult:
    values: List[str] = field(default_factory=list)
    raw_row_count: int = 0
    duplicates_skipped: int = 0
    sheet_name: str = ""
    header_row: int = 0
    header_col: int = 0


def _cell_text(value: Any) -> str:
    """Normalize cell text: collapse whitespace/newlines/NBSP, then strip."""
    if value is None:
        return ""
    s = str(value).replace("\u00a0", " ")
    s = " ".join(s.split())
    return s.strip()


def find_header_cell(
    worksheet,
    header_name: str,
    max_rows: int = HEADER_SCAN_MAX_ROWS,
    max_cols: int = HEADER_SCAN_MAX_COLS,
) -> Optional[Tuple[int, int]]:
    """
    Locate the best header cell for header_name (case-insensitive).

    For ambiguous short headers like "NAME", prefer the match whose column
    below has the most non-blank values (data column scoring).
    """
    target = header_name.strip().upper()
    accepted_headers = HEADER_ALIASES.get(target, frozenset({target}))
    max_r = min(worksheet.max_row or 1, max_rows)
    max_c = min(worksheet.max_column or 1, max_cols)

    candidates: List[Tuple[int, int, int]] = []  # (data_count, row, col)
    for row in range(1, max_r + 1):
        for col in range(1, max_c + 1):
            text = _cell_text(worksheet.cell(row=row, column=col).value)
            if text.upper() not in accepted_headers:
                continue
            data_count = 0
            for r in range(row + 1, min(row + 200, (worksheet.max_row or row) + 1)):
                if _cell_text(worksheet.cell(row=r, column=col).value):
                    data_count += 1
            candidates.append((data_count, row, col))

    if not candidates:
        # Fallback: full-sheet scan if window missed the header
        max_r = worksheet.max_row or 1
        max_c = worksheet.max_column or 1
        for row in range(1, max_r + 1):
            for col in range(1, max_c + 1):
                text = _cell_text(worksheet.cell(row=row, column=col).value)
                if text.upper() in accepted_headers:
                    data_count = 0
                    for r in range(row + 1, min(row + 200, max_r + 1)):
                        if _cell_text(worksheet.cell(row=r, column=col).value):
                            data_count += 1
                    candidates.append((data_count, row, col))

    if not candidates:
        return None

    # Prefer richest data column; tie-break by earliest row/col
    candidates.sort(key=lambda t: (-t[0], t[1], t[2]))
    _, row, col = candidates[0]
    return row, col


def _resolve_worksheet(wb, header_name: str):
    """Scan every sheet and choose the richest matching data column."""
    best = None
    best_score = -1
    best_preference = len(PREFERRED_SHEET_NAMES) + 1
    for ws in wb.worksheets:
        loc = find_header_cell(ws, header_name)
        if not loc:
            continue
        header_row, header_col = loc
        score = 0
        for r in range(header_row + 1, min(header_row + 500, (ws.max_row or header_row) + 1)):
            if _cell_text(ws.cell(row=r, column=header_col).value):
                score += 1
        preference = (
            PREFERRED_SHEET_NAMES.index(ws.title)
            if ws.title in PREFERRED_SHEET_NAMES
            else len(PREFERRED_SHEET_NAMES)
        )
        if score > best_score or (
            score == best_score and preference < best_preference
        ):
            best_score = score
            best_preference = preference
            best = ws

    return best


def extract_column_values(
    workbook_path: Path | str,
    header_name: str,
    sheet_index: int | None = None,
) -> ExtractionResult:
    """
    Open a workbook, locate `header_name`, and return unique trimmed values.

    Dedup is case-insensitive for engineering tags; original casing of the
    first occurrence is preserved for reporting.
    """
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    suffix = path.suffix.lower()
    if suffix == ".xls":
        raise ValueError(
            f'Legacy .xls format is not supported for "{path.name}". '
            "Please save/export as .xlsx and retry."
        )

    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if not wb.worksheets:
            raise ValueError(f"Workbook has no worksheets: {path.name}")

        if sheet_index is not None:
            if sheet_index < 0 or sheet_index >= len(wb.worksheets):
                raise ValueError(
                    f"Worksheet index {sheet_index} out of range for {path.name}"
                )
            ws = wb.worksheets[sheet_index]
        else:
            ws = _resolve_worksheet(wb, header_name)

        if ws is None:
            raise ValueError(
                f'Could not find header "{header_name}" in any worksheet of "{path.name}".'
            )

        header = find_header_cell(ws, header_name)
        if header is None:
            raise ValueError(
                f'Could not find header "{header_name}" on sheet "{ws.title}" in "{path.name}".'
            )

        header_row, header_col = header
        values: List[str] = []
        seen_upper: set = set()
        raw_count = 0
        duplicates = 0

        for row in range(header_row + 1, (ws.max_row or header_row) + 1):
            text = _cell_text(ws.cell(row=row, column=header_col).value)
            if not text:
                continue
            raw_count += 1
            key = text.upper()
            if key in seen_upper:
                duplicates += 1
                continue
            seen_upper.add(key)
            values.append(text)

        result = ExtractionResult(
            values=values,
            raw_row_count=raw_count,
            duplicates_skipped=duplicates,
            sheet_name=ws.title,
            header_row=header_row,
            header_col=header_col,
        )
        logger.info(
            'Extracted "%s" from %s sheet="%s" header=(%s,%s) '
            "raw=%s unique=%s duplicates=%s",
            header_name,
            path.name,
            result.sheet_name,
            result.header_row,
            result.header_col,
            result.raw_row_count,
            len(result.values),
            result.duplicates_skipped,
        )
        return result
    finally:
        wb.close()


def extract_device_tags(workbook_path: Path | str) -> List[str]:
    """Extract unique Device Tag values from Worksheet 1 workbook."""
    return extract_column_values(workbook_path, "Device Tag").values


def extract_names(workbook_path: Path | str) -> List[str]:
    """Extract unique NAME values from Worksheet 2 workbook."""
    return extract_column_values(workbook_path, "NAME").values


def extract_device_tags_detailed(workbook_path: Path | str) -> ExtractionResult:
    return extract_column_values(workbook_path, "Device Tag")


def extract_names_detailed(workbook_path: Path | str) -> ExtractionResult:
    return extract_column_values(workbook_path, "NAME")


def extract_engineering_device_tags_detailed(
    workbook_path: Path | str,
) -> ExtractionResult:
    """Extract the exact ``$(DEVICETAG)`` engineering import column."""
    return extract_column_values(workbook_path, "$(DEVICETAG)")
