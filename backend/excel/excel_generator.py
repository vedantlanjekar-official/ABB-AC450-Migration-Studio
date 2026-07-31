from pathlib import Path
from typing import Dict, List, Any
import openpyxl
from openpyxl.utils import get_column_letter
from backend.core.logging import get_logger
from backend.excel.design import build_db_excel_design
from backend.excel.header_postprocessor import (
    DB_HEADER_MAPPING,
    rename_export_headers,
)

class ExcelGenerator:
    """
    Generates Valmet-compatible formatted Excel workbooks from grouped DB element datasets.
    Creates dynamic worksheets (one per element type), applies professional industrial styling,
    zebra striping, header formatting, and auto-adjusted column widths.
    """

    def __init__(self, job_id: str = None):
        self.logger = get_logger(job_id)

    def generate_workbook(self, mapped_sheets: Dict[str, List[Dict[str, Any]]], output_path: Path) -> List[str]:
        """
        Creates an Excel workbook at output_path.
        
        Args:
            mapped_sheets: Dict[sheet_name, list of row dicts]
            output_path: Path where .xlsx file will be saved
            
        Returns:
            List of generated sheet names
        """
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active

        generated_sheets = []

        design = build_db_excel_design()

        for sheet_name, rows in mapped_sheets.items():
            if not rows:
                continue

            ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet title max length 31
            generated_sheets.append(sheet_name)

            # Determine all unique columns preserving order
            columns = list(rows[0].keys())

            # Write header row
            ws.append(columns)
            for col_num in range(1, len(columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = design.header_fill
                cell.font = design.header_font
                cell.alignment = design.center_align

            # Write data rows
            for row_idx, row_data in enumerate(rows, start=2):
                row_values = [row_data.get(col, "") for col in columns]
                ws.append(row_values)
                
                fill_color = (
                    design.zebra_fill if row_idx % 2 == 0 else design.white_fill
                )
                for col_num in range(1, len(columns) + 1):
                    raw_val = ws.cell(row=row_idx, column=col_num).value
                    if isinstance(raw_val, str):
                        clean_str = raw_val.lstrip('=+@').strip()
                        if clean_str.startswith('-') and len(clean_str) > 1 and clean_str[1].isalpha():
                            clean_str = clean_str[1:].strip()
                        cell = ws.cell(row=row_idx, column=col_num, value=clean_str)
                        cell.data_type = 's'
                    else:
                        cell = ws.cell(row=row_idx, column=col_num)

                    cell.fill = fill_color
                    cell.font = design.cell_font
                    cell.border = design.thin_border
                    
                    # Align center for short codes / numbers / tags, left for descriptions
                    val_str = str(cell.value or "")
                    if len(val_str) < 15 and not " " in val_str:
                        cell.alignment = design.center_align
                    else:
                        cell.alignment = design.left_align

            # Auto-fit column widths
            for col_idx, col_name in enumerate(columns, start=1):
                max_len = len(str(col_name))
                for row_num in range(2, len(rows) + 2):
                    cell_val = str(ws.cell(row=row_num, column=col_idx).value or "")
                    if len(cell_val) > max_len:
                        max_len = len(cell_val)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

        # Remove default sheet if sheets were generated
        if generated_sheets and default_sheet in wb.worksheets:
            wb.remove(default_sheet)

        # Final post-processing step: transform only displayed export headers.
        rename_export_headers(wb, DB_HEADER_MAPPING, header_row=1)
        wb.save(output_path)
        self.logger.info(f"Generated Excel workbook at {output_path} with sheets: {generated_sheets}")
        return generated_sheets
