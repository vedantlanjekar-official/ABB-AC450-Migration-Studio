"""Shared Excel design used by DB, PC, and comparison exports."""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class ExcelDesign:
    header_fill: PatternFill
    header_font: Font
    cell_font: Font
    zebra_fill: PatternFill
    white_fill: PatternFill
    thin_border: Border
    center_align: Alignment
    left_align: Alignment


def build_db_excel_design() -> ExcelDesign:
    """Return fresh style objects matching the DB Element workbook."""
    thin_side = Side(style="thin", color="E2E8F0")
    return ExcelDesign(
        header_fill=PatternFill(
            start_color="1E293B",
            end_color="1E293B",
            fill_type="solid",
        ),
        header_font=Font(
            name="Calibri",
            size=11,
            bold=True,
            color="FFFFFF",
        ),
        cell_font=Font(name="Calibri", size=10, color="000000"),
        zebra_fill=PatternFill(
            start_color="F8FAFC",
            end_color="F8FAFC",
            fill_type="solid",
        ),
        white_fill=PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        ),
        thin_border=Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        ),
        center_align=Alignment(horizontal="center", vertical="center"),
        left_align=Alignment(horizontal="left", vertical="center"),
    )
