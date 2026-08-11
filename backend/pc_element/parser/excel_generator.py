"""
excel_generator.py - Stage 12: Excel Generator for PC Element Engineering I/O Lists.

Writes a consolidated I/O worksheet plus an independent Function Block Summary
sheet. Row order for I/O is preserved as provided (clubbing + output formatting
happen upstream). This layer does not reorder or recalculate extracted
engineering values.

I_O_List columns:
  Sr. No. | Loop Tag | Description | Device Tag |
  AI | AO | DI | DO | AI800_ | AO800_ | DI800_ | DO800_ |
  Slot/Card | Channel

Function Block Summary columns:
  Functional Block | Total Count
"""

from typing import List, Any, Optional, Mapping
import os
import openpyxl
from openpyxl.utils import get_column_letter
from backend.excel.design import build_db_excel_design
from backend.pc_element.parser.validator import EngineeringIO
from backend.pc_element.parser.category_mapper import (
    CATEGORY_INDICATOR_COLUMNS,
    build_category_indicator_values,
)
from backend.pc_element.parser.function_block_extractor import (
    SUPPORTED_FUNCTION_BLOCKS,
    function_block_summary_rows,
)
from backend.excel.header_postprocessor import (
    PC_HEADER_MAPPING,
    rename_export_headers,
)

FUNCTION_BLOCK_SUMMARY_SHEET = "Function Block Summary"


def sanitize_cell_value(val: Any) -> Any:
    """Sanitizes cell values to prevent openpyxl from treating strings starting with = or - as Excel formulas."""
    if val is None:
        return ""
    if isinstance(val, str):
        s = val.lstrip('=+@').strip()
        if s.startswith('-') and len(s) > 1 and s[1].isalpha():
            s = s[1:].strip()
        return s
    return val


class ExcelGenerator:
    """Generates Valmet-compatible Excel workbooks for PC Element I/O lists."""

    # Columns that replace the former single Category field
    _CATEGORY_COLS = CATEGORY_INDICATOR_COLUMNS  # 5..12 after Device Tag
    _CENTER_COLS = {1, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14}  # Sr + indicators + Slot/Card + Channel

    @classmethod
    def generate_excel(
        cls,
        io_objects: List[EngineeringIO],
        output_path: str,
        function_block_counts: Optional[Mapping[str, int]] = None,
    ) -> str:
        """
        Generate a styled workbook at output_path with I_O_List and
        Function Block Summary sheets.

        Preserves the caller-provided I/O row order (Loop Tag clubbing /
        formatting must be applied before this call). Returns the filepath.
        """
        ordered_objects = list(io_objects)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "I_O_List"
        ws.views.sheetView[0].showGridLines = True

        design = build_db_excel_design()

        headers = [
            "Sr. No.",
            "Loop Tag",
            "Description",
            "Device Tag",
            *CATEGORY_INDICATOR_COLUMNS,
            "Slot/Card",
            "Channel",
        ]
        start_row = 1
        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header_text)
            cell.fill = design.header_fill
            cell.font = design.header_font
            cell.alignment = design.center_align
            cell.border = design.thin_border
        ws.row_dimensions[start_row].height = 26

        current_row = start_row + 1
        for sr_no, obj in enumerate(ordered_objects, start=1):
            channel_val = obj.channel_number if obj.channel_number > 0 else ""
            indicators = build_category_indicator_values(obj.category)
            row_data = [
                sr_no,
                obj.loop_tag,
                obj.description or "",
                obj.device_tag,
                *[indicators[col] for col in CATEGORY_INDICATOR_COLUMNS],
                obj.card_number,
                channel_val,
            ]

            fill_to_apply = (
                design.zebra_fill if sr_no % 2 == 0 else design.white_fill
            )

            for col_idx, val in enumerate(row_data, start=1):
                if col_idx == 1 or col_idx >= 5:
                    clean_val = val if val != "" else ""
                else:
                    clean_val = sanitize_cell_value(val)
                    if clean_val is None:
                        clean_val = ""
                cell = ws.cell(row=current_row, column=col_idx, value=clean_val)
                cell.font = design.cell_font
                cell.border = design.thin_border

                if isinstance(clean_val, str):
                    cell.data_type = 's'

                cell.fill = fill_to_apply

                if col_idx in cls._CENTER_COLS:
                    cell.alignment = design.center_align
                else:
                    cell.alignment = design.left_align

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < start_row:
                    continue
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Keep indicator columns compact
            header_val = ws.cell(row=start_row, column=col[0].column).value or ""
            if header_val in CATEGORY_INDICATOR_COLUMNS:
                width = max(max_len + 2, 8)
            else:
                width = max(max_len + 4, 12)
            ws.column_dimensions[col_letter].width = min(width, 60)

        if os.path.dirname(output_path):
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
        # Transform only I_O_List export headers before appending the summary sheet.
        rename_export_headers(wb, PC_HEADER_MAPPING, header_row=start_row)
        cls._write_function_block_summary(wb, function_block_counts, design)
        wb.save(output_path)
        return output_path

    @classmethod
    def _write_function_block_summary(
        cls,
        wb: openpyxl.Workbook,
        function_block_counts: Optional[Mapping[str, int]],
        design,
    ) -> None:
        """Append the Function Block Summary worksheet (independent of I/O)."""
        ws = wb.create_sheet(FUNCTION_BLOCK_SUMMARY_SHEET)
        ws.views.sheetView[0].showGridLines = True

        headers = ["Functional Block", "Total Count"]
        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.fill = design.header_fill
            cell.font = design.header_font
            cell.alignment = design.center_align
            cell.border = design.thin_border
        ws.row_dimensions[1].height = 26

        rows = function_block_summary_rows(
            function_block_counts,
            block_names=SUPPORTED_FUNCTION_BLOCKS,
        )
        for sr_no, row in enumerate(rows, start=1):
            fill_to_apply = (
                design.zebra_fill if sr_no % 2 == 0 else design.white_fill
            )
            values = [row["Functional Block"], row["Total Count"]]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=sr_no + 1, column=col_idx, value=val)
                cell.font = design.cell_font
                cell.border = design.thin_border
                cell.fill = fill_to_apply
                if col_idx == 1:
                    cell.alignment = design.left_align
                    if isinstance(val, str):
                        cell.data_type = "s"
                else:
                    cell.alignment = design.center_align
            ws.row_dimensions[sr_no + 1].height = 20

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 14
