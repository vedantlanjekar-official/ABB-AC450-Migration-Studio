import pytest
import openpyxl
from pathlib import Path
from backend.pc_parser.grammar_parser import PCGrammarParser
from backend.pc_parser.pdf_reader import PCLineRecord
from backend.pc_parser.description_mapper import PCDescriptionMapper
from backend.pc_parser.duplicate_checker import PCDuplicateChecker
from backend.pc_parser.excel_generator import PCExcelGenerator
from backend.pc_parser.parser_service import PCParserService
from backend.models.pc_element import PCElement

def test_pc_grammar_parser_extraction():
    parser = PCGrammarParser("test_pc_grammar")

    # Test 1: AI1.1/940LC391.MV
    ref1 = parser.parse_reference("AI1.1/940LC391.MV")
    assert ref1 is not None
    assert ref1["Category"] == "AI"
    assert ref1["Card"] == "1"
    assert ref1["Channel"] == "1"
    assert ref1["DeviceTag"] == "940LC391.MV"
    assert ref1["LoopTag"] == "940LC391"

    # Test 2: AO800_3.6/945FC400.OUT
    ref2 = parser.parse_reference("AO800_3.6/945FC400.OUT")
    assert ref2 is not None
    assert ref2["Category"] == "AO800"
    assert ref2["Card"] == "3"
    assert ref2["Channel"] == "6"
    assert ref2["DeviceTag"] == "945FC400.OUT"
    assert ref2["LoopTag"] == "945FC400"

    # Test 3: DI1.14/940M03M1.RUN
    ref3 = parser.parse_reference("  DI1.14/940M03M1.RUN  ")
    assert ref3 is not None
    assert ref3["Category"] == "DI"
    assert ref3["Card"] == "1"
    assert ref3["Channel"] == "14"
    assert ref3["DeviceTag"] == "940M03M1.RUN"
    assert ref3["LoopTag"] == "940M03M1"

    # Test 4: DO4.5/946M22M2.STOP
    ref4 = parser.parse_reference("DO4.5/946M22M2.STOP")
    assert ref4 is not None
    assert ref4["Category"] == "DO"
    assert ref4["Card"] == "4"
    assert ref4["Channel"] == "5"
    assert ref4["DeviceTag"] == "946M22M2.STOP"
    assert ref4["LoopTag"] == "946M22M2"

def test_pc_description_mapper():
    records = [
        PCLineRecord(page_number=1, line_number=1, text="AI1.2"),
        PCLineRecord(page_number=1, line_number=2, text="AI1.2/940LC391.MV"),
        PCLineRecord(page_number=1, line_number=3, text="PURE WATER TANK LEVEL"),
    ]

    grammar_parser = PCGrammarParser("test_desc")
    detected_ref = grammar_parser.parse_reference(records[1].text)
    assert detected_ref is not None

    mapper = PCDescriptionMapper("test_desc")
    results = mapper.attach_descriptions([(detected_ref, records[1])], records)

    assert len(results) == 1
    assert results[0]["Description"] == "PURE WATER TANK LEVEL"

def test_pc_duplicate_checker():
    elements = [
        PCElement(category="AI", card_number="1", channel_number="1", device_tag="940LC391.MV", loop_tag="940LC391", description="Level 1"),
        PCElement(category="AI", card_number="1", channel_number="1", device_tag="940LC391.MV", loop_tag="940LC391", description="Level 1 Dup"),
        PCElement(category="AO", card_number="3", channel_number="6", device_tag="945FC400.OUT", loop_tag="945FC400", description="Flow 1")
    ]

    checker = PCDuplicateChecker("test_dedup")
    deduped, dup_count = checker.deduplicate_and_number(elements)

    assert len(deduped) == 2
    assert dup_count == 1
    assert deduped[0].sr_no == 1
    assert deduped[1].sr_no == 2

def test_pc_excel_generator(tmp_path):
    elements = [
        PCElement(category="AO", card_number="3", channel_number="6", device_tag="945FC400.OUT", loop_tag="945FC400", description="Flow Out"),
        PCElement(category="AI", card_number="1", channel_number="14", device_tag="940M03M1.MV", loop_tag="940M03M1", description="Temp 1"),
        PCElement(category="AI", card_number="1", channel_number="2", device_tag="940LC391.MV", loop_tag="940LC391", description="Level 1"),
    ]

    out_file = tmp_path / "pc_valmet_export.xlsx"
    generator = PCExcelGenerator("test_excel")
    generator.generate_excel(elements, out_file)

    assert out_file.exists()
    wb = openpyxl.load_workbook(out_file)
    assert "Valmet PC Export" in wb.sheetnames

    ws = wb["Valmet PC Export"]
    assert ws.cell(row=1, column=1).value == "Sr No"
    assert ws.cell(row=1, column=2).value == "Loop Tag"

    # Sorted by Category (AI -> AO) then Card -> Channel
    # Row 2 (1st data row): AI, Card 1, Channel 2 (AI1.2)
    assert ws.cell(row=2, column=5).value == "AI"
    assert ws.cell(row=2, column=6).value == "1"
    assert ws.cell(row=2, column=7).value == "2"

    # Row 3 (2nd data row): AI, Card 1, Channel 14 (AI1.14)
    assert ws.cell(row=3, column=5).value == "AI"
    assert ws.cell(row=3, column=6).value == "1"
    assert ws.cell(row=3, column=7).value == "14"

    # Row 4 (3rd data row): AO, Card 3, Channel 6 (AO3.6)
    assert ws.cell(row=4, column=5).value == "AO"

def test_pc_parser_service_full_pipeline():
    records = [
        PCLineRecord(page_number=1, line_number=1, text="AI1.1/940LC391.MV"),
        PCLineRecord(page_number=1, line_number=2, text="PURE WATER TANK LEVEL"),
        PCLineRecord(page_number=1, line_number=3, text="AO3.6/945FC400.OUT"),
        PCLineRecord(page_number=1, line_number=4, text="COOLING TOWER FLOW CONTROL"),
        PCLineRecord(page_number=1, line_number=5, text="DI1.14/940M03M1.RUN"),
        PCLineRecord(page_number=1, line_number=6, text="HYPO TWR PUMP1 RUNNING"),
        PCLineRecord(page_number=1, line_number=7, text="DO4.5/946M22M2.STOP"),
        PCLineRecord(page_number=1, line_number=8, text="ST. CL03 FD PUMP2 STOP"),
    ]

    service = PCParserService("test_service")
    deduped, stats, warnings = service.parse_line_records(records)

    assert len(deduped) == 4
    assert stats.engineering_references_found == 4
    assert stats.ai_count == 1
    assert stats.ao_count == 1
    assert stats.di_count == 1
    assert stats.do_count == 1
    assert stats.duplicate_records == 0
    assert stats.missing_descriptions == 0

    tags = [e.device_tag for e in deduped]
    assert "940LC391.MV" in tags
    assert "945FC400.OUT" in tags
    assert "940M03M1.RUN" in tags
    assert "946M22M2.STOP" in tags
