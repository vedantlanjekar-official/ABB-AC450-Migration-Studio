import pytest
import openpyxl
from pathlib import Path
from backend.excel.excel_generator import ExcelGenerator
from backend.parser.pdf_text_extractor import PDFTextExtractor
from backend.parser.parser_service import ParserService
from backend.mapper.element_mapper import ElementMapper
from backend.mapper.record_clubber import RecordClubber

def test_excel_workbook_generation(tmp_path):
    mapped_sheets = {
        "AI": [
            {"Tag": "AI1.1", "Index": "1.1", "NAME": "PRESS_01", "UNIT": "BAR", "RANGEMAX": 100.0},
            {"Tag": "AI1.2", "Index": "1.2", "NAME": "PRESS_02", "UNIT": "BAR", "RANGEMAX": 200.0}
        ],
        "DI": [
            {"Tag": "DI3.1", "Index": "3.1", "NAME": "LIMIT_01", "INV": 0}
        ]
    }
    
    out_file = tmp_path / "test_out.xlsx"
    generator = ExcelGenerator("test_job_3")
    sheets = generator.generate_workbook(mapped_sheets, out_file)
    
    assert out_file.exists()
    assert "AI" in sheets
    assert "DI" in sheets
    
    wb = openpyxl.load_workbook(out_file)
    assert "AI" in wb.sheetnames
    assert "DI" in wb.sheetnames
    
    ws_ai = wb["AI"]
    assert ws_ai.cell(row=1, column=1).value == "Tag"
    assert ws_ai.cell(row=2, column=1).value == "AI1.1"
    assert ws_ai.cell(row=2, column=3).value == "PRESS_01"

def test_excel_single_clubbed_worksheet(tmp_path):
    """DB export must produce one consolidated Clubbed_IO worksheet."""
    sample_pdf = Path(__file__).resolve().parent.parent.parent / "examples" / "sample_ac450_db.pdf"
    extractor = PDFTextExtractor("test_excel_defaults")
    pages = extractor.extract_text_pages(sample_pdf)

    parser = ParserService("test_excel_defaults")
    elements, stats, _ = parser.parse_document_pages(pages, "sample_ac450_db.pdf")

    clubbed = RecordClubber("test_excel_defaults").club_elements(elements)
    mapper = ElementMapper("test_excel_defaults")
    clubbed_rows = mapper.map_clubbed(clubbed)

    categories = {r["Category"] for r in clubbed_rows}
    assert categories.issubset(
        {"AI", "AO", "DI", "DO", "AI800", "AO800", "DI800", "DO800"}
    )
    assert "PIDCON" not in categories

    out_excel = tmp_path / "valmet_defaults_export.xlsx"
    generator = ExcelGenerator("test_excel_defaults")
    generator.generate_workbook({"Clubbed_IO": clubbed_rows}, out_excel)

    wb = openpyxl.load_workbook(out_excel)
    assert wb.sheetnames == ["Clubbed_IO"]
    ws = wb["Clubbed_IO"]

    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    assert "Category" in headers
    assert "TYPE" in headers
    assert "SCANT" in headers
    assert "DEC" in headers
    assert "ERR_TR" in headers

    type_col = headers.index("TYPE") + 1
    scant_col = headers.index("SCANT") + 1
    dec_col = headers.index("DEC") + 1

    ai1_1_row = next(
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=headers.index("Tag") + 1).value == "AI1.1"
    )
    # AI1.1 inherited TYPE/SCANT/DEC from defaults
    assert ws.cell(row=ai1_1_row, column=type_col).value == "ANALOG_INPUT"
    assert ws.cell(row=ai1_1_row, column=scant_col).value == "1s"
    assert ws.cell(row=ai1_1_row, column=dec_col).value == 2

def test_excel_no_blank_cells_for_inherited_defaults():
    from backend.models.db_element import DBElement

    # AI1.1 has explicit TYPE, AI1.2 omits TYPE
    elements = [
        DBElement(
            tag="AI1.1",
            element_type="AI",
            element_index="1.1",
            parameters={"NAME": "SENS_1", "TYPE": "ANALOG_INPUT", "SCANT": "1s"}
        ),
        DBElement(
            tag="AI1.2",
            element_type="AI",
            element_index="1.2",
            parameters={"NAME": "SENS_2"}
        )
    ]

    mapper = ElementMapper("test_defaults_filling")
    rows = mapper.map_clubbed(elements)
    ai_rows = [r for r in rows if r["Category"] == "AI"]

    row_1 = next(r for r in ai_rows if r["Tag"] == "AI1.1")
    row_2 = next(r for r in ai_rows if r["Tag"] == "AI1.2")

    # Row 2 MUST NOT have empty string for TYPE or SCANT, it must take the sheet default value!
    assert row_2["TYPE"] == "ANALOG_INPUT"
    assert row_2["SCANT"] == "1s"
