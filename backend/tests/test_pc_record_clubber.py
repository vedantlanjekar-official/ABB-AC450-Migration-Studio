"""Tests for PC Element Loop Tag record clubbing and output formatting."""

from backend.pc_element.parser.validator import EngineeringIO
from backend.pc_element.parser.record_clubber import RecordClubber
from backend.pc_element.parser.output_formatter import OutputFormatter
from backend.pc_element.parser.excel_generator import ExcelGenerator
from backend.mapper.record_clubber import derive_loop_tag, derive_name_suffix
import openpyxl


def _io(
    category: str,
    device_tag: str,
    loop_tag: str = "",
    card: int = 1,
    channel: int = 1,
    io_family: str = "",
) -> EngineeringIO:
    """Build a minimal EngineeringIO for clubbing tests."""
    cat = category.upper()
    family = io_family or (f"{cat}_" if cat.endswith("800") else cat)
    # Normalize: AI800 category uses io_family AI800_
    if cat in {"AI800", "AO800", "DI800", "DO800"} and not io_family:
        family = f"{cat}_"
    resolved_loop = loop_tag or derive_loop_tag(device_tag)
    return EngineeringIO(
        io_family=family,
        io_type=cat.replace("800", "") if "800" in cat else cat,
        category=cat,
        card_number=card,
        channel_number=channel,
        loop_tag=resolved_loop,
        device_tag=device_tag,
        source_reference=f"={family}{card}.{channel}/{device_tag}",
    )


def test_derive_loop_tag_shared_with_db():
    assert derive_loop_tag("940FQ390.MV") == "940FQ390"
    assert derive_loop_tag("940FQ390.OUT") == "940FQ390"
    assert derive_loop_tag("940XV101.RUN") == "940XV101"
    assert derive_loop_tag("940XV101.SV1") == "940XV101"
    assert derive_name_suffix("940XV101.GSO") == "GSO"


def test_ai_ao_clubbing_order():
    """AI → AO for the same Loop Tag; unpaired AI kept without placeholder."""
    elements = [
        _io("AO", "940FQ390.OUT", card=2),
        _io("AI", "940FQ390.MV", card=1),
        _io("AI", "941TI100.MV", card=3),  # unpaired AI
        _io("AO", "ZZZ999.OUT", card=9),   # different loop
    ]
    clubbed = RecordClubber("test").club_elements(elements)
    devices = [e.device_tag for e in clubbed]
    categories = [e.category for e in clubbed]

    assert devices[0] == "940FQ390.MV"
    assert devices[1] == "940FQ390.OUT"
    assert devices[2] == "941TI100.MV"
    assert devices[3] == "ZZZ999.OUT"
    assert categories[:2] == ["AI", "AO"]
    assert len(clubbed) == 4


def test_do_di_clubbing_order():
    """DO → DI for the same Loop Tag."""
    elements = [
        _io("DI", "940XV101.RUN"),
        _io("DO", "940XV101.OUT"),
    ]
    clubbed = RecordClubber("test").club_elements(elements)
    assert [e.category for e in clubbed] == ["DO", "DI"]
    assert [e.device_tag for e in clubbed] == ["940XV101.OUT", "940XV101.RUN"]


def test_family_sections_ai_ao_before_do_di():
    """All AI–AO clubs complete before any DO–DI."""
    elements = [
        _io("DI", "100XV001.RUN"),
        _io("DO", "100XV001.OUT"),
        _io("AO", "940FQ390.OUT"),
        _io("AI", "940FQ390.MV"),
    ]
    clubbed = RecordClubber("test").club_elements(elements)
    assert [e.category for e in clubbed] == ["AI", "AO", "DO", "DI"]
    assert [e.device_tag for e in clubbed] == [
        "940FQ390.MV",
        "940FQ390.OUT",
        "100XV001.OUT",
        "100XV001.RUN",
    ]


def test_ai800_ao800_and_do800_di800_section_order():
    """AI800–AO800 section before DO800–DI800 section."""
    elements = [
        _io("DI800", "100XV101.RUN"),
        _io("DO800", "100XV101.OUT"),
        _io("AO800", "940FQ390.OUT"),
        _io("AI800", "940FQ390.MV"),
    ]
    clubbed = RecordClubber("test").club_elements(elements)
    assert [e.category for e in clubbed] == ["AI800", "AO800", "DO800", "DI800"]


def test_full_section_sequence():
    """Global order: AI–AO → DO–DI → AI800–AO800 → DO800–DI800."""
    elements = [
        _io("DI800", "Z99.RUN"),
        _io("DO", "B10.OUT"),
        _io("AI800", "C20.MV"),
        _io("AI", "A01.MV"),
        _io("AO", "A01.OUT"),
        _io("DI", "B10.RUN"),
        _io("AO800", "C20.OUT"),
        _io("DO800", "Z99.OUT"),
    ]
    clubbed = RecordClubber("test").club_elements(elements)
    assert [e.category for e in clubbed] == [
        "AI", "AO",
        "DO", "DI",
        "AI800", "AO800",
        "DO800", "DI800",
    ]


def test_valve_suffix_order_sv1_gso_gsc():
    elements = [
        _io("DI", "940XV200.GSC", channel=3),
        _io("DO", "940XV200.SV1", channel=1),
        _io("DO", "940XV200.GSO", channel=2),
    ]
    clubbed = RecordClubber("test").club_elements(elements)
    assert [e.device_tag for e in clubbed] == [
        "940XV200.SV1",
        "940XV200.GSO",
        "940XV200.GSC",
    ]


def test_output_formatter_preserves_pairs_and_section_order():
    """Formatter reorders jumbled clubs into engineering sequence without data loss."""
    elements = [
        _io("DI800", "Z99.RUN"),
        _io("AO", "A01.OUT"),
        _io("DO800", "Z99.OUT"),
        _io("AI", "A01.MV"),
        _io("DI", "B10.RUN"),
        _io("DO", "B10.OUT"),
        _io("AO800", "C20.OUT"),
        _io("AI800", "C20.MV"),
    ]
    clubbed = RecordClubber("test").club_elements(elements)
    formatted = OutputFormatter("test").format_clubbed_elements(clubbed)

    assert [e.category for e in formatted] == [
        "AI", "AO",
        "DO", "DI",
        "AI800", "AO800",
        "DO800", "DI800",
    ]
    assert len(formatted) == len(elements)
    # Extracted values unchanged
    by_device = {e.device_tag: e for e in formatted}
    assert by_device["A01.MV"].category == "AI"
    assert by_device["A01.OUT"].category == "AO"
    assert by_device["B10.OUT"].loop_tag == "B10"


def test_no_placeholder_for_unpaired_records():
    elements = [
        _io("AI", "940FQ390.MV"),
        # no AO counterpart
    ]
    clubbed = RecordClubber("test").club_elements(elements)
    formatted = OutputFormatter("test").format_clubbed_elements(clubbed)
    assert len(formatted) == 1
    assert formatted[0].device_tag == "940FQ390.MV"


def test_excel_preserves_clubbed_order(tmp_path):
    """Single worksheet; rows follow AI→AO then DO→DI clubbing order."""
    elements = [
        _io("DI", "940XV101.RUN", card=3, channel=1),
        _io("DO", "940XV101.OUT", card=4, channel=1),
        _io("AO", "940FQ390.OUT", card=2, channel=1),
        _io("AI", "940FQ390.MV", card=1, channel=1),
    ]
    clubbed = RecordClubber("test").club_elements(elements)
    formatted = OutputFormatter("test").format_clubbed_elements(clubbed)

    out_file = str(tmp_path / "pc_clubbed.xlsx")
    ExcelGenerator.generate_excel(formatted, out_file)
    wb = openpyxl.load_workbook(out_file)
    assert wb.sheetnames == ["I_O_List"]
    ws = wb["I_O_List"]

    categories_ai = [ws.cell(row=r, column=5).value for r in range(2, 6)]
    categories_ao = [ws.cell(row=r, column=6).value for r in range(2, 6)]
    categories_di = [ws.cell(row=r, column=7).value for r in range(2, 6)]
    categories_do = [ws.cell(row=r, column=8).value for r in range(2, 6)]
    devices = [ws.cell(row=r, column=4).value for r in range(2, 6)]
    loops = [ws.cell(row=r, column=2).value for r in range(2, 6)]

    # Indicator columns replace Category: AI→AO→DO→DI order
    assert categories_ai == [1, None, None, None]
    assert categories_ao == [None, 1, None, None]
    assert categories_do == [None, None, 1, None]
    assert categories_di == [None, None, None, 1]
    assert devices == [
        "940FQ390.MV",
        "940FQ390.OUT",
        "940XV101.OUT",
        "940XV101.RUN",
    ]
    assert loops == ["940FQ390", "940FQ390", "940XV101", "940XV101"]
