"""AAX PC Element reader tests (redesigned soft + hardwired extraction)."""

from pathlib import Path
import tempfile

import openpyxl
import pytest

from backend.pc_element.parser.aax_reader import AaxReader
from backend.pc_element.parser.function_block_extractor import count_function_blocks
from backend.pc_element.parser.grammar_parser import GrammarParser
from backend.pc_element.parser.io_reference_detector import IOReferenceDetector
from backend.pc_element.parser.parser_service import PCParserService

# The AAX reference corpus lives at repository-root / ".AXX Data".
# The old hard-coded Downloads path is preserved as a fallback so historical
# working copies still resolve; the CI-friendly repo-relative path is preferred.
_REPO_ROOT = Path(__file__).resolve().parents[2]
_REPO_DATASET = _REPO_ROOT / ".AXX Data"
_LEGACY_DATASET = Path(r"c:\Users\admin\Downloads\PM2MP2\PCDATA")
SAMPLE_AAX_DIR = _REPO_DATASET if _REPO_DATASET.exists() else _LEGACY_DATASET
FIXTURE_AAX = Path(__file__).resolve().parent / "fixtures" / "sample_pc.aax"

ALL_SAMPLES = [
    "23JA0301.AAX",
    "23JA0401.AAX",
    "23JA0501.AAX",
    "23JA1001.AAX",
    "23JA1101.AAX",
    "23JA1201.AAX",
    "23JA1301.AAX",
    "23JA1401.AAX",
    "23JA1501.AAX",
    "23JA1601.AAX",
    "23JA1701.AAX",
    "23JA1801.AAX",
    "23JA1901.AAX",
    "23JA2001.AAX",
    "23JA2101.AAX",
    "N220301.AAX",
    "N220401.AAX",
    "N220501.AAX",
    "N220601.AAX",
    "N221001.AAX",
    "N221101.AAX",
    "N221201.AAX",
    "N221301.AAX",
    "N221401.AAX",
    "N221501.AAX",
    "N221601.AAX",
    "N221701.AAX",
    "N221801.AAX",
    "N221901.AAX",
    "N222001.AAX",
    "N222101.AAX",
]


def _write_fixture(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        """BEGIN PC16
HEADER
     IEC_DocNo      22-PC16
     IEC_Title2     PM2 \\ Test Area
     IEC_Project1   Carter Holt Harvey Tissue

PCD-PAGE    1
\"82PIC972: Steam Box Supply\"

PC16.31.22.1.110 PIDCON (1,1,1,1,1,0)
          :DBINST             =82PIC972
          :MV                 =AI7.10
          :ALCBLK             =AI7.10:ERR
          :OUTP

PC16.31.22.1.111 MOVE (R,2)
          :21                 =82PIC972:POUT
          :22                 =AO3.9

PCD-PAGE    2
\"82PIC740: Soft MV Example\"

PC16.32.12.1.110 PIDCON (1,1,1,1,1,0)
          :DBINST             =82PIC740
          :MV                 =82PIC740.MV
          :PARAM1             =82PIC740:PARAM1

PC16.32.12.1.111 MOVE (R,1)
          :22                 =82PIC740.OUT

PC16.31.21.1 MOVE (B,2)
          :21                 =DI2.23:BLOCKED
          :22                 =DI1.31:BLOCKED

PC16.31.21.16 OR (2)
          :20                 =DO3.19

PC16.33.12.1.110 VALVECON (0,1)
          :DBINST             =82HSV632
PC16.33.12.1.111 MOVE (B,1)
          :21                 =82HSV632.SV

PC16.32.12.1.210 MOTCON (0,0,1,1,1,0)
          :DBINST             =82M001

PC16.40.1.1 OR (2)
          :21                 =82LIC428:BLK_D
          :22                 =82FIC480:POUT

END PC16
""",        encoding="utf-8",
    )


@pytest.fixture(scope="module", autouse=True)
def ensure_fixture():
    _write_fixture(FIXTURE_AAX)


def test_aax_reader_synthesizes_hardwired_and_soft():
    pages = AaxReader(str(FIXTURE_AAX)).read_all_pages()
    joined = "\n".join(p.text for p in pages)
    assert "PIDCON (" in joined
    assert "=AI7.10/" in joined
    assert "=AO3.9/" in joined
    assert "=DI2.23/" in joined
    assert "=DO3.19/" in joined
    # Soft MV / OUT / SV must become grammar candidates (card 0)
    assert "=AI0.0/82PIC740.MV" in joined or "/82PIC740.MV" in joined
    assert "/82PIC740.OUT" in joined or "/82PIC740.OUT" in joined.upper()
    assert "/82HSV632.SV" in joined.upper()
    assert "/82LIC428.BLK_D" in joined.upper()
    assert "/82FIC480.POUT" in joined.upper()

    candidates = []
    for p in pages:
        candidates.extend(
            IOReferenceDetector.detect_candidates_in_page(
                p.text, p.raw_lines, text_layers=p.text_layers
            )
        )
    refs = []
    for c in candidates:
        refs.extend(GrammarParser.parse_all_references(c))
    tags = {r.device_tag for r in refs}
    assert any(t.startswith("82PIC972") for t in tags)
    assert "82PIC740.MV" in tags
    assert "82HSV632.SV" in tags


def test_aax_fixture_full_pipeline_has_io_and_fb():
    with tempfile.TemporaryDirectory() as tmp:
        result = PCParserService(
            file_path=str(FIXTURE_AAX),
            job_id="test_aax_pipe",
            output_dir=tmp,
        ).execute_pipeline()

        assert not result.errors, result.errors
        assert result.total_io_found >= 6
        assert result.ai_count >= 1
        assert result.ao_count >= 1
        assert result.excel_file_path and Path(result.excel_file_path).exists()

        wb = openpyxl.load_workbook(result.excel_file_path)
        assert "I_O_List" in wb.sheetnames
        assert "Function Block Summary" in wb.sheetnames
        fb_rows = {
            ws[0]: ws[1]
            for ws in wb["Function Block Summary"].iter_rows(
                min_row=2, max_col=2, values_only=True
            )
        }
        assert fb_rows.get("PIDCON", 0) >= 2
        assert fb_rows.get("MOTCON", 0) >= 1
        assert fb_rows.get("VALVECON", 0) >= 1


def test_soft_only_sample_extracts_mv_tags():
    path = SAMPLE_AAX_DIR / "N221601.AAX"
    if not path.exists():
        pytest.skip("sample missing")
    with tempfile.TemporaryDirectory() as tmp:
        result = PCParserService(str(path), "n221601", tmp).execute_pipeline()
    assert not result.errors
    assert result.total_io_found >= 10
    assert result.ai_count >= 5


def test_colon_soft_sample_extracts_blk_and_pout():
    path = SAMPLE_AAX_DIR / "23JA1001.AAX"
    if not path.exists():
        pytest.skip("sample missing")
    with tempfile.TemporaryDirectory() as tmp:
        result = PCParserService(str(path), "ja1001", tmp).execute_pipeline()
    assert not result.errors
    assert result.total_io_found >= 30


def test_hardwired_sample_preserves_card_channel():
    path = SAMPLE_AAX_DIR / "23JA1601.AAX"
    if not path.exists():
        pytest.skip("sample missing")
    with tempfile.TemporaryDirectory() as tmp:
        result = PCParserService(str(path), "ja1601", tmp).execute_pipeline()
        assert not result.errors
        assert result.total_io_found >= 19
        wb = openpyxl.load_workbook(result.excel_file_path)
        ws = wb["I_O_List"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        slot_idx = headers.index("Slot/Card") if "Slot/Card" in headers else None
        assert slot_idx is not None
        cards = [
            row[slot_idx]
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[slot_idx]
        ]
        assert any(isinstance(c, int) and c > 0 for c in cards)


def _excel_tag_slot_channel(path: Path):
    import tempfile
    from backend.pc_element.parser.parser_service import PCParserService
    with tempfile.TemporaryDirectory() as tmp:
        result = PCParserService(str(path), "addr", tmp).execute_pipeline()
        wb = openpyxl.load_workbook(result.excel_file_path)
        ws = wb["I_O_List"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        tag_idx = headers.index("$(DEVICETAG)") if "$(DEVICETAG)" in headers else 3
        slot_idx = headers.index("Slot/Card")
        ch_idx = headers.index("Channel")
        rows = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            tag = (r[tag_idx] or "").upper()
            if tag:
                rows[tag] = (r[slot_idx], r[ch_idx])
        return rows, result


def test_pidcon_mv_and_move_pout_resolve_card_channel():
    path = SAMPLE_AAX_DIR / "23JA1601.AAX"
    if not path.exists():
        pytest.skip("sample missing")
    rows, _ = _excel_tag_slot_channel(path)
    assert rows.get("82PIC972.MV") == (7, 10)
    pout = rows.get("82PIC972.POUT") or rows.get("82PIC972.OUT")
    assert pout == (3, 9)
    assert rows.get("82HSV974.CLS") == (3, 20)
    assert rows.get("82M103.M2") == (1, 32)
    assert rows.get("82M103.SO1") == (1, 25)


def test_pcu_ioaddr_channel_bound_to_speed_tag():
    path = SAMPLE_AAX_DIR / "23JA1801.AAX"
    if not path.exists():
        pytest.skip("sample missing")
    rows, _ = _excel_tag_slot_channel(path)
    assert rows.get("82M140.SPEEDMV") == (192, 1)
    assert rows.get("82M140.PULSEOUT") == (192, 1)
    assert rows.get("82M136.SPEEDMV") == (192, 2)
    assert rows.get("82SIA958.SPEEDMV") == (193, 1)


def test_preview_shows_extracted_slot_channel_beside_device_tag():
    """Results preview must include resolved card/channel, not hide them after indicators."""
    path = SAMPLE_AAX_DIR / "23JA1601.AAX"
    if not path.exists():
        pytest.skip("sample missing")
    with tempfile.TemporaryDirectory() as tmp:
        result = PCParserService(str(path), "preview_addr", tmp).execute_pipeline()
    assert result.preview_data
    keys = list(result.preview_data[0].keys())
    device_idx = keys.index("Device Tag")
    assert keys[device_idx + 1] == "Slot/Card"
    assert keys[device_idx + 2] == "Channel"
    by_tag = {
        str(row.get("Device Tag") or "").upper(): row
        for row in result.preview_data
    }
    pic = by_tag.get("82PIC972.MV")
    assert pic is not None
    assert pic.get("Slot/Card") == 7
    assert pic.get("Channel") == 10


def test_aax_excel_fills_page_title_descriptions():
    path = SAMPLE_AAX_DIR / "23JA0501.AAX"
    if not path.exists():
        pytest.skip("sample missing")
    pages = AaxReader(str(path)).read_all_pages()
    joined_head = "\n".join(pages[0].text.splitlines()[:8])
    assert "122F124A" in joined_head
    assert "Communications" in joined_head

    with tempfile.TemporaryDirectory() as tmp:
        result = PCParserService(str(path), "ja0501_desc", tmp).execute_pipeline()
        wb = openpyxl.load_workbook(result.excel_file_path)
        ws = wb["I_O_List"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        desc_idx = headers.index("$(NAME_40)") if "$(NAME_40)" in headers else 2
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 20
        filled = sum(1 for r in rows if (r[desc_idx] or "").strip())
        assert filled / len(rows) >= 0.8


@pytest.mark.parametrize("filename", ALL_SAMPLES)
def test_every_sample_aax_produces_excel(filename: str):
    path = SAMPLE_AAX_DIR / filename
    if not path.exists():
        pytest.skip(f"missing {filename}")
    pages = AaxReader(str(path)).read_all_pages()
    assert pages
    fb = count_function_blocks([p.text for p in pages])
    assert set(fb) >= {"PIDCON", "MOTCON", "VALVECON", "MANSTN"}

    with tempfile.TemporaryDirectory() as tmp:
        result = PCParserService(str(path), f"s_{path.stem}", tmp).execute_pipeline()
        assert not result.errors, result.errors
        assert result.excel_file_path
        assert Path(result.excel_file_path).exists()
        wb = openpyxl.load_workbook(result.excel_file_path)
        assert "I_O_List" in wb.sheetnames
        assert "Function Block Summary" in wb.sheetnames


def test_multi_aax_job_excel_contains_every_file():
    """Downloading a multi-AAX PC job must write all files into one workbook."""
    f1 = SAMPLE_AAX_DIR / "23JA0401.AAX"
    f2 = SAMPLE_AAX_DIR / "23JA0501.AAX"
    if not f1.exists() or not f2.exists():
        pytest.skip("sample missing")

    import shutil
    import uuid
    from backend.core.config import settings
    from backend.services.conversion_service import ConversionService
    from backend.services.job_manager import job_store

    with tempfile.TemporaryDirectory() as tmp:
        n1 = len(
            PCParserService(str(f1), "multi_a", tmp).execute_pipeline().exported_objects
        )
        n2 = len(
            PCParserService(str(f2), "multi_b", tmp).execute_pipeline().exported_objects
        )

    job_id = f"test_multi_aax_{uuid.uuid4().hex[:8]}"
    upload_dir = settings.UPLOAD_DIR / job_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy(f1, upload_dir / f1.name)
    shutil.copy(f2, upload_dir / f2.name)
    job_store.create_job(job_id, [f1.name, f2.name])
    ConversionService(job_id)._run_pc_conversion_pipeline()

    job = job_store.get_job(job_id)
    assert job["status"] == "completed"
    excel_path = Path(job["excel_file_path"])
    assert excel_path.exists()
    assert excel_path.name == "PC_Element_IO_List.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    rows = list(wb["I_O_List"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == n1 + n2
    assert job["total_objects"] == n1 + n2
    tags = {(r[3] or "").upper() for r in rows if r[3]}
    assert any(t.startswith("122F124") for t in tags)
    assert "Function Block Summary" in wb.sheetnames
