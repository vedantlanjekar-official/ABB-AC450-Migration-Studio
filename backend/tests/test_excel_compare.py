"""Tests for Excel Comparison & Validation module."""

from pathlib import Path
import openpyxl

from backend.excel_compare.column_extractor import (
    extract_device_tags,
    extract_names,
    find_header_cell,
)
from backend.excel_compare.comparator import ExcelComparator, UnmatchedTag
from backend.excel_compare.report_generator import ComparisonReportGenerator


def _make_ws1(path: Path, tags: list[str], header_row: int = 4, header_col: int = 4):
    """Create an arbitrary workbook containing $(DEVICETAG)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=header_row, column=header_col, value="$(DEVICETAG)")
    for i, tag in enumerate(tags):
        ws.cell(row=header_row + 1 + i, column=header_col, value=tag)
    wb.save(path)


def _make_ws2(path: Path, names: list[str], header_row: int = 1, header_col: int = 4):
    """Create a second arbitrary workbook containing $(DEVICETAG)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=header_row, column=header_col, value="$(DEVICETAG)")
    for i, name in enumerate(names):
        ws.cell(row=header_row + 1 + i, column=header_col, value=name)
    wb.save(path)


def test_find_header_device_tag_dynamic(tmp_path):
    path = tmp_path / "ws1.xlsx"
    _make_ws1(path, ["940FQ390.MV"], header_row=4, header_col=4)
    wb = openpyxl.load_workbook(path)
    loc = find_header_cell(wb.active, "Device Tag")
    assert loc == (4, 4)


def test_find_header_name_dynamic(tmp_path):
    path = tmp_path / "ws2.xlsx"
    _make_ws2(path, ["940FQ390.MV"], header_row=1, header_col=4)
    wb = openpyxl.load_workbook(path)
    loc = find_header_cell(wb.active, "NAME")
    assert loc == (1, 4)


def test_extract_ignores_blanks_and_duplicates(tmp_path):
    path = tmp_path / "ws1.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["D4"] = "Device Tag"
    ws["D5"] = "940FQ390.MV"
    ws["D6"] = "  "
    ws["D7"] = "940FQ390.MV"  # duplicate
    ws["D8"] = " 940PT210.OUT "
    wb.save(path)

    tags = extract_device_tags(path)
    assert tags == ["940FQ390.MV", "940PT210.OUT"]


def test_extractors_accept_engineering_device_tag_header(tmp_path):
    path = tmp_path / "engineering_headers.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet["D4"] = "$(DEVICETAG)"
    worksheet["D5"] = "940FQ390.MV"
    workbook.save(path)

    assert extract_device_tags(path) == ["940FQ390.MV"]
    assert extract_names(path) == ["940FQ390.MV"]


def test_comparison_matched_and_unmatched(tmp_path):
    ws1 = tmp_path / "file1.xlsx"
    ws2 = tmp_path / "file2.xlsx"
    _make_ws1(
        ws1,
        [
            "940FQ390.MV",
            "940FQ390.OUT",
            "940PT210.OUT",
            "940XV101.RUN",
            "MISSING.TAG",
        ],
    )
    _make_ws2(
        ws2,
        [
            "940FQ390.MV",
            "940FQ390.OUT",
            "940PT210.OUT",
            "940XV101.RUN",
            "EXTRA.IN.WS2",
        ],
    )

    result = ExcelComparator().compare(ws1, ws2)
    assert result.worksheet1_records == 5
    assert result.worksheet2_records == 5
    assert result.matched_records == 4
    assert result.unmatched_records == 2
    assert result.unmatched_tags == ["MISSING.TAG", "EXTRA.IN.WS2"]
    assert result.unmatched_in_worksheet1 == ["MISSING.TAG"]
    assert result.unmatched_in_worksheet2 == ["EXTRA.IN.WS2"]
    assert result.unmatched_items[0].source_file == "file1.xlsx"
    assert result.unmatched_items[1].source_file == "file2.xlsx"


def test_shuffled_order_all_matched_set_based(tmp_path):
    """
    Records in completely different order must still all match.
    Proves comparison is value/set based — never Row N ↔ Row N.
    """
    ws1 = tmp_path / "file1.xlsx"
    ws2 = tmp_path / "file2.xlsx"

    # Worksheet 1 order
    _make_ws1(
        ws1,
        [
            "940FQ390.MV",
            "940LT210.MV",
            "940XV101.RUN",
            "940PT120.MV",
        ],
    )
    # Worksheet 2 — same tags, completely shuffled
    _make_ws2(
        ws2,
        [
            "940PT120.MV",
            "940XV101.RUN",
            "940FQ390.MV",
            "940LT210.MV",
        ],
    )

    result = ExcelComparator().compare(ws1, ws2)
    assert result.worksheet1_records == 4
    assert result.worksheet2_records == 4
    assert result.matched_records == 4
    assert result.unmatched_records == 0
    assert result.unmatched_tags == []
    assert set(result.matched_tags) == {
        "940FQ390.MV",
        "940LT210.MV",
        "940XV101.RUN",
        "940PT120.MV",
    }


def test_shuffled_order_partial_unmatched(tmp_path):
    """Shuffled NAME column still finds matches; only missing values are unmatched."""
    ws1 = tmp_path / "file1.xlsx"
    ws2 = tmp_path / "file2.xlsx"
    _make_ws1(
        ws1,
        [
            "940FQ390.MV",
            "940LT210.MV",
            "940XV101.RUN",
            "ONLY.IN.WS1",
        ],
    )
    _make_ws2(
        ws2,
        [
            "940XV101.RUN",
            "ONLY.IN.WS2",
            "940FQ390.MV",
            "940LT210.MV",
        ],
    )

    result = ExcelComparator().compare(ws1, ws2)
    assert result.matched_records == 3
    assert result.unmatched_records == 2
    assert result.unmatched_tags == ["ONLY.IN.WS1", "ONLY.IN.WS2"]


def test_exact_match_is_case_insensitive(tmp_path):
    """Engineering tags match case-insensitively after trim."""
    ws1 = tmp_path / "file1.xlsx"
    ws2 = tmp_path / "file2.xlsx"
    _make_ws1(ws1, ["940fq390.mv"])
    _make_ws2(ws2, ["940FQ390.MV"])
    result = ExcelComparator().compare(ws1, ws2)
    assert result.matched_records == 1
    assert result.unmatched_records == 0


def test_multi_sheet_picks_data_sheet(tmp_path):
    """Cover sheet first must not hide Device Tag / NAME data sheets."""
    ws1 = tmp_path / "file1.xlsx"
    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Instructions"
    data = wb.create_sheet("I_O_List")
    data["D4"] = "$(DEVICETAG)"
    data["D5"] = "940FQ390.MV"
    data["D6"] = "940LT210.MV"
    wb.save(ws1)

    ws2 = tmp_path / "file2.xlsx"
    wb2 = openpyxl.Workbook()
    cover2 = wb2.active
    cover2.title = "Notes"
    cover2["A1"] = "$(DEVICETAG)"  # false-positive header with no data
    data2 = wb2.create_sheet("Clubbed_IO")
    data2["D1"] = "$(DEVICETAG)"
    data2["D2"] = "940LT210.MV"
    data2["D3"] = "940FQ390.MV"
    wb2.save(ws2)

    result = ExcelComparator().compare(ws1, ws2)
    assert result.matched_records == 2
    assert result.unmatched_records == 0
    assert result.worksheet1_sheet == "I_O_List"
    assert result.worksheet2_sheet == "Clubbed_IO"


def test_trimmed_exact_match(tmp_path):
    """Leading/trailing spaces are trimmed before exact comparison."""
    ws1 = tmp_path / "file1.xlsx"
    ws2 = tmp_path / "file2.xlsx"
    _make_ws1(ws1, ["  940FQ390.MV  "])
    _make_ws2(ws2, ["940FQ390.MV"])
    result = ExcelComparator().compare(ws1, ws2)
    assert result.matched_records == 1
    assert result.unmatched_records == 0


def test_report_generator_sheets(tmp_path):
    ws1 = tmp_path / "file1.xlsx"
    ws2 = tmp_path / "file2.xlsx"
    _make_ws1(ws1, ["A.MV", "B.OUT", "C.RUN"])
    _make_ws2(ws2, ["A.MV", "B.OUT", "D.EXTRA"])
    result = ExcelComparator().compare(ws1, ws2)

    out = tmp_path / "Comparison_Report.xlsx"
    ComparisonReportGenerator().generate(result, out)
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Summary", "Unmatched Records"]

    summary = wb["Summary"]
    # Metrics start at row 6
    metrics = {
        summary.cell(row=r, column=1).value: summary.cell(row=r, column=2).value
        for r in range(6, 13)
        if summary.cell(row=r, column=1).value
    }
    assert metrics["Worksheet 1 Records"] == 3
    assert metrics["Worksheet 2 Records"] == 3
    assert metrics["Matched Records"] == 2
    assert metrics["Total Unmatched Records"] == 2
    assert metrics["Only in Excel 1"] == 1
    assert metrics["Only in Excel 2"] == 1
    assert summary["A5"].fill.fgColor.rgb.endswith("1E293B")
    assert summary["A5"].font.name == "Calibri"
    assert summary["A5"].font.bold is True
    assert summary["A6"].fill.fgColor.rgb.endswith("FFFFFF")
    assert summary["A7"].fill.fgColor.rgb.endswith("F8FAFC")

    unmatched = wb["Unmatched Records"]
    assert unmatched.cell(row=4, column=2).value == "C.RUN"
    assert unmatched.cell(row=4, column=3).value == "file1.xlsx"
    assert unmatched.cell(row=4, column=4).value == "file2.xlsx"
    assert unmatched.cell(row=5, column=2).value == "D.EXTRA"
    assert unmatched.cell(row=5, column=3).value == "file2.xlsx"
    assert unmatched.cell(row=5, column=4).value == "file1.xlsx"
    assert unmatched["A3"].fill.fgColor.rgb.endswith("1E293B")
    assert unmatched["A3"].font.name == "Calibri"
    assert unmatched["A4"].font.name == "Calibri"


def test_preview_payload():
    from backend.excel_compare.comparator import ComparisonResult

    result = ComparisonResult(
        worksheet1_records=2,
        worksheet2_records=1,
        matched_records=1,
        unmatched_records=1,
        unmatched_tags=["X.MV"],
        unmatched_in_worksheet1=["X.MV"],
        unmatched_items=[
            UnmatchedTag(
                tag="X.MV",
                source_file="db.xlsx",
                source_sheet="Clubbed_IO",
                missing_from_file="pc.xlsx",
            )
        ],
    )
    preview = ComparisonReportGenerator.build_preview(result)
    assert "Summary" in preview
    assert "Unmatched Records" in preview
    assert preview["Unmatched Records"][0]["$(DEVICETAG)"] == "X.MV"
    assert preview["Unmatched Records"][0]["Source Excel"] == "db.xlsx"
    assert "Source Sheet" not in preview["Unmatched Records"][0]
