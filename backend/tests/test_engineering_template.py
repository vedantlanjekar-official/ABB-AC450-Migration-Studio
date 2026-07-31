import io
import time
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from backend.engineering_template import (
    EngineeringTemplateGenerator,
    TEMPLATE_COLUMNS,
)
from backend.engineering_template.generator import SourceRecord
from backend.main import app
from backend.mapper.category_mapper import CATEGORY_INDICATOR_COLUMNS


client = TestClient(app)


def _append_indicator_row(worksheet, headers, values_by_header):
    row = [""] * len(headers)
    for header, value in values_by_header.items():
        row[headers.index(header)] = value
    worksheet.append(row)


def _build_db_clubbed_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Clubbed_IO"
    headers = [
        "Tag",
        "$(TAG)",
        "$(DEVICETAG)",
        "$(NAME_40)",
        "$(DEVICETAG:UNIT)",
        "$(DEVICETAG:MIN)",
        "$(DEVICETAG:MAX)",
        "$(PACKAGE)",
        "Process Area ID",
        "$(EXE)",
        "$(CTRLROOM)",
        "$(ALGROUP)",
        *CATEGORY_INDICATOR_COLUMNS,
    ]
    worksheet.append(headers)

    rows = [
        {
            "Tag": "AI1.1",
            "$(TAG)": "949TI001",
            "$(DEVICETAG)": "949TI001.MV",
            "$(NAME_40)": "Temp In",
            "$(DEVICETAG:UNIT)": "C",
            "$(DEVICETAG:MIN)": 0,
            "$(DEVICETAG:MAX)": 100,
            "$(PACKAGE)": "PKG-A",
            "Process Area ID": "AREA-1",
            "$(EXE)": "EXE1",
            "$(CTRLROOM)": "CR1",
            "$(ALGROUP)": "AL1",
            "AI": 1,
        },
        {
            "Tag": "AO1.1",
            "$(TAG)": "949TI001",
            "$(DEVICETAG)": "949TI001.OUT",
            "$(NAME_40)": "Temp Out",
            "$(DEVICETAG:UNIT)": "C",
            "$(DEVICETAG:MIN)": 0,
            "$(DEVICETAG:MAX)": 100,
            "AO": 1,
        },
        {
            "Tag": "DO1.1",
            "$(TAG)": "949XV002",
            "$(DEVICETAG)": "949XV002.CMD",
            "$(NAME_40)": "Valve Cmd",
            "DO": 1,
        },
        {
            "Tag": "DI1.1",
            "$(TAG)": "949XV002",
            "$(DEVICETAG)": "949XV002.FB",
            "$(NAME_40)": "Valve FB",
            "DI": 1,
        },
        {
            "Tag": "AI800_1.1",
            "$(TAG)": "940AI800",
            "$(DEVICETAG)": "940AI800.MV",
            "$(NAME_40)": "AI800 In",
            "AI800_": 1,
        },
        {
            "Tag": "AO800_1.1",
            "$(TAG)": "940AI800",
            "$(DEVICETAG)": "940AI800.OUT",
            "$(NAME_40)": "AO800 Out",
            "AO800_": 1,
        },
        {
            "Tag": "DO800_1.1",
            "$(TAG)": "940DO800",
            "$(DEVICETAG)": "940DO800.CMD",
            "$(NAME_40)": "DO800 Cmd",
            "DO800_": 1,
        },
        {
            "Tag": "DI800_1.1",
            "$(TAG)": "940DO800",
            "$(DEVICETAG)": "940DO800.FB",
            "$(NAME_40)": "DI800 FB",
            "DI800_": 1,
        },
        {
            "Tag": "AI1.2",
            "$(TAG)": "949TI099",
            "$(DEVICETAG)": "949TI099.MV",
            "$(NAME_40)": "Singleton AI",
            "AI": 1,
        },
        {
            "Tag": "AI1.3",
            "$(TAG)": "LOOP-A",
            "$(DEVICETAG)": "LOOP-A.MV",
            "$(NAME_40)": "Mismatch AI",
            "AI": 1,
        },
        {
            "Tag": "AO1.2",
            "$(TAG)": "LOOP-B",
            "$(DEVICETAG)": "LOOP-B.OUT",
            "$(NAME_40)": "Mismatch AO",
            "AO": 1,
        },
    ]
    for values in rows:
        _append_indicator_row(worksheet, headers, values)

    workbook.save(path)


def _build_pc_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "I_O_List"
    # PC headers are on row 4
    worksheet.append(["PC Element Export"])
    worksheet.append(["Generated"])
    worksheet.append([])
    headers = [
        "Sr. No.",
        "$(TAG)",
        "Description",
        "$(DEVICETAG)",
        *CATEGORY_INDICATOR_COLUMNS,
    ]
    worksheet.append(headers)
    # Indicator columns: AI AO DI DO AI800_ AO800_ DI800_ DO800_
    worksheet.append(
        [1, "940M01", "Motor Run", "940M01.RUN", "", "", "", "", "", "", 1, ""]
    )
    worksheet.append(
        [2, "940M01", "Motor Cmd", "940M01.CMD", "", "", "", "", "", "", "", 1]
    )
    workbook.save(path)


def test_generator_pairs_all_families_and_preserves_order(tmp_path):
    source = tmp_path / "clubbed_db.xlsx"
    output = tmp_path / "ABB_Engineering_Template.xlsx"
    _build_db_clubbed_workbook(source)

    result = EngineeringTemplateGenerator().generate(source, output)

    assert result.source_records == 11
    assert result.paired_clubs == 4
    assert result.singleton_rows == 3
    assert len(result.template_rows) == 7
    assert result.generated_sheets == ["Engineering_Template"]
    assert any("different Loop Tags" in warning for warning in result.warnings)

    rows = result.template_rows
    # AI + AO → AI slot1, AO slot2
    assert rows[0]["$(CARDTYPE1)"] == "AI"
    assert rows[0]["$(DEVICETAG1)"] == "949TI001.MV"
    assert rows[0]["$(CARDTYPE2)"] == "AO"
    assert rows[0]["$(DEVICETAG2)"] == "949TI001.OUT"
    assert rows[0]["$(TAG)"] == "949TI001"
    assert rows[0]["$(NAME40_1)"] == "Temp In"
    assert rows[0]["$(DEVICETAG1:UNIT)"] == "C"
    assert rows[0]["$(DEVICETAG1:MIN)"] == 0
    assert rows[0]["$(DEVICETAG1:MAX)"] == 100
    assert rows[0]["$(PACKAGE)"] == "PKG-A"
    assert rows[0]["Process Area ID"] == "AREA-1"
    assert rows[0]["$(EXE)"] == "EXE1"
    assert rows[0]["$(CTRLROOM)"] == "CR1"
    assert rows[0]["$(ALGROUP)"] == "AL1"

    # DO + DI adjacent → DI slot1, DO slot2
    assert rows[1]["$(CARDTYPE1)"] == "DI"
    assert rows[1]["$(DEVICETAG1)"] == "949XV002.FB"
    assert rows[1]["$(CARDTYPE2)"] == "DO"
    assert rows[1]["$(DEVICETAG2)"] == "949XV002.CMD"

    # AI800_ + AO800_
    assert rows[2]["$(CARDTYPE1)"] == "AI800_"
    assert rows[2]["$(CARDTYPE2)"] == "AO800_"

    # DO800_ + DI800_ → DI800_ first
    assert rows[3]["$(CARDTYPE1)"] == "DI800_"
    assert rows[3]["$(DEVICETAG1)"] == "940DO800.FB"
    assert rows[3]["$(CARDTYPE2)"] == "DO800_"
    assert rows[3]["$(DEVICETAG2)"] == "940DO800.CMD"

    # Singleton
    assert rows[4]["$(CARDTYPE1)"] == "AI"
    assert rows[4]["$(DEVICETAG1)"] == "949TI099.MV"
    assert rows[4]["$(CARDTYPE2)"] == ""
    assert rows[4]["$(DEVICETAG2)"] == ""

    # Mismatched adjacent Loop Tags stay separate
    assert rows[5]["$(DEVICETAG1)"] == "LOOP-A.MV"
    assert rows[5]["$(CARDTYPE2)"] == ""
    assert rows[6]["$(DEVICETAG1)"] == "LOOP-B.OUT"
    assert rows[6]["$(CARDTYPE2)"] == ""

    workbook = openpyxl.load_workbook(output)
    try:
        sheet = workbook["Engineering_Template"]
        headers = [sheet.cell(1, col).value for col in range(1, len(TEMPLATE_COLUMNS) + 1)]
        assert headers == list(TEMPLATE_COLUMNS)
        assert sheet.cell(2, TEMPLATE_COLUMNS.index("$(DEVICETAG1)") + 1).value == "949TI001.MV"
        assert sheet.cell(2, TEMPLATE_COLUMNS.index("$(DEVICETAG2)") + 1).value == "949TI001.OUT"
        # Shared DB navy header styling
        header_fill = sheet.cell(1, 1).fill.fgColor
        assert header_fill is not None
        assert str(header_fill.rgb or header_fill.theme).upper() != ""
    finally:
        workbook.close()


def test_generator_reads_pc_schema_and_digital_input_first(tmp_path):
    source = tmp_path / "pc.xlsx"
    output = tmp_path / "template.xlsx"
    _build_pc_workbook(source)

    result = EngineeringTemplateGenerator().generate(source, output)

    assert result.source_records == 2
    assert result.paired_clubs == 1
    assert result.singleton_rows == 0
    row = result.template_rows[0]
    assert row["$(CARDTYPE1)"] == "DI800_"
    assert row["$(DEVICETAG1)"] == "940M01.RUN"
    assert row["$(CARDTYPE2)"] == "DO800_"
    assert row["$(DEVICETAG2)"] == "940M01.CMD"
    assert row["$(NAME40_1)"] == "Motor Run"
    assert row["$(TAG)"] == "940M01"
    # Optional package fields absent → blank
    assert row["$(PACKAGE)"] == ""
    assert row["Process Area ID"] == ""


def test_legacy_db_headers_and_category_column(tmp_path):
    source = tmp_path / "legacy.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Clubbed_IO"
    worksheet.append(["NAME", "Loop Tag", "DESCR", "UNIT", "Category"])
    worksheet.append(["TAG1.MV", "LOOP1", "Desc 1", "bar", "AI"])
    worksheet.append(["TAG1.OUT", "LOOP1", "Desc 2", "bar", "AO"])
    workbook.save(source)

    result = EngineeringTemplateGenerator().generate(source, tmp_path / "out.xlsx")
    assert result.paired_clubs == 1
    row = result.template_rows[0]
    assert row["$(DEVICETAG1)"] == "TAG1.MV"
    assert row["$(DEVICETAG2)"] == "TAG1.OUT"
    assert row["$(DEVICETAG1:UNIT)"] == "bar"


def test_optional_blanks_when_headers_missing():
    generator = EngineeringTemplateGenerator()
    rows, paired, singletons, warnings = generator.map_records(
        [
            SourceRecord(
                category="AI",
                device_tag="A.MV",
                loop_tag="L1",
                description="Only AI",
                source_row=2,
            )
        ]
    )
    assert paired == 0
    assert singletons == 1
    assert warnings == []
    assert rows[0]["$(PACKAGE)"] == ""
    assert rows[0]["$(EXE)"] == ""
    assert rows[0]["$(CARDTYPE2)"] == ""


def test_rejects_legacy_xls(tmp_path):
    source = tmp_path / "legacy.xls"
    source.write_bytes(b"not-a-real-xls")
    with pytest.raises(ValueError, match=r"\.xls"):
        EngineeringTemplateGenerator().generate(source, tmp_path / "out.xlsx")


def test_engineering_template_api_flow(tmp_path):
    source = tmp_path / "clubbed_db.xlsx"
    _build_db_clubbed_workbook(source)

    with source.open("rb") as file_handle:
        upload = client.post(
            "/api/upload",
            files={
                "files": (
                    source.name,
                    file_handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert upload.status_code == 200
    job_id = upload.json()["job_id"]

    process = client.post(
        "/api/process",
        json={"job_id": job_id, "conversion_type": "ENG_TEMPLATE"},
    )
    assert process.status_code == 200
    assert process.json()["conversion_type"] == "ENG_TEMPLATE"

    status = None
    for _ in range(40):
        response = client.get(f"/api/status/{job_id}")
        assert response.status_code == 200
        status = response.json()
        if status["status"] in {"completed", "failed"}:
            break
        time.sleep(0.1)

    assert status is not None
    assert status["status"] == "completed", status.get("errors")
    assert status["conversion_type"] == "ENG_TEMPLATE"
    assert status["total_objects"] == 7
    assert status["matched_records"] == 4
    assert status["unmatched_records"] == 3
    assert status["generated_sheets"] == ["Engineering_Template"]
    assert "Engineering_Template" in status["preview_data"]

    download = client.get(f"/api/download/{job_id}")
    assert download.status_code == 200
    assert "ABB_Engineering_Template.xlsx" in download.headers["content-disposition"]
    exported = openpyxl.load_workbook(io.BytesIO(download.content), data_only=True)
    try:
        sheet = exported["Engineering_Template"]
        headers = [sheet.cell(1, col).value for col in range(1, len(TEMPLATE_COLUMNS) + 1)]
        assert headers == list(TEMPLATE_COLUMNS)
        assert sheet.cell(2, TEMPLATE_COLUMNS.index("$(CARDTYPE1)") + 1).value == "AI"
        assert sheet.cell(2, TEMPLATE_COLUMNS.index("$(CARDTYPE2)") + 1).value == "AO"
        assert sheet.cell(3, TEMPLATE_COLUMNS.index("$(CARDTYPE1)") + 1).value == "DI"
        assert sheet.cell(3, TEMPLATE_COLUMNS.index("$(CARDTYPE2)") + 1).value == "DO"
    finally:
        exported.close()
