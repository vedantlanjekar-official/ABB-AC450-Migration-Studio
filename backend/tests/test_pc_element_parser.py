"""
test_pc_element_parser.py - Unit tests for ABB AC450 PC Element Hardwired I/O Parser.
"""

import pytest
import os
import openpyxl
from backend.pc_element.parser.grammar_parser import GrammarParser
from backend.pc_element.parser.duplicate_detector import DuplicateDetector
from backend.pc_element.parser.validator import Validator, EngineeringIO
from backend.pc_element.parser.excel_generator import (
    ExcelGenerator,
    FUNCTION_BLOCK_SUMMARY_SHEET,
)
from backend.pc_element.parser.io_reference_detector import IOReferenceDetector
from backend.pc_element.parser.function_block_extractor import (
    count_function_blocks,
    function_block_summary_rows,
    SUPPORTED_FUNCTION_BLOCKS,
)


def test_user_examples_extraction():
    """Test all explicit user example formats from the request prompt."""

    ref1 = GrammarParser.parse_reference("=AI1.1/940LC391.MV")
    assert ref1 is not None
    assert ref1.category == "AI"
    assert ref1.card_number == 1
    assert ref1.channel_number == 1
    assert ref1.loop_tag == "940LC391"
    assert ref1.device_tag == "940LC391.MV"

    ref2 = GrammarParser.parse_reference("=AO2.3/945FC400.OUT")
    assert ref2 is not None
    assert ref2.category == "AO"
    assert ref2.device_tag == "945FC400.OUT"

    ref3 = GrammarParser.parse_reference("=DI5.12/940M03M1.RUN")
    assert ref3 is not None
    assert ref3.category == "DI"

    ref4 = GrammarParser.parse_reference("=DO8.6/946M22M2.STOP")
    assert ref4 is not None
    assert ref4.category == "DO"

    ref5 = GrammarParser.parse_reference("=AI800_1.1/940LC391.MV")
    assert ref5 is not None
    assert ref5.category == "AI800"
    assert ref5.io_family == "AI800_"
    assert ref5.loop_tag == "940LC391"
    assert ref5.device_tag == "940LC391.MV"

    ref6 = GrammarParser.parse_reference("DO800_1.16/82M073.MSTR")
    assert ref6 is not None
    assert ref6.category == "DO800"
    assert ref6.io_type == "DO"

    ref7 = GrammarParser.parse_reference("-AI800_2.1/M49M021.CURR")
    assert ref7 is not None
    assert ref7.device_tag == "M49M021.CURR"
    assert ref7.loop_tag == "M49M021"


def test_channel_port_err_format():
    """CARD.CHANNEL:TERMINAL/TAG:ERR — Device Tag drops the :ERR attribute."""
    r = GrammarParser.parse_reference("=AI800_22.5:22/M49FI1201.MV:ERR")
    assert r is not None
    assert r.category == "AI800"
    assert r.card_number == 22
    assert r.channel_number == 5
    assert r.device_tag == "M49FI1201.MV"
    assert r.loop_tag == "M49FI1201"

    r2 = GrammarParser.parse_reference("=AI800_2.1:22/M49M021.CURR:ERR")
    assert r2 is not None
    assert r2.card_number == 2
    assert r2.channel_number == 1
    assert r2.device_tag == "M49M021.CURR"
    assert r2.loop_tag == "M49M021"


def test_mv_and_err_collapse_to_same_device_tag():
    """MV and MV:ERR normalize to the same Device Tag and dedupe when address matches."""
    refs = [
        GrammarParser.parse_reference("=AI800_22.5/M49FI1201.MV"),
        GrammarParser.parse_reference("=AI800_22.5:22/M49FI1201.MV:ERR"),
    ]
    unique, dups = DuplicateDetector.deduplicate_references(refs)
    assert len(unique) == 1
    assert dups == 1
    assert unique[0].device_tag == "M49FI1201.MV"


def test_device_tag_strips_colon_attributes():
    """Device Tag keeps only the engineering tag — colon attributes are discarded."""
    cases = [
        ("=AI800_22.5:22/M49FI1201.MV:ERR", "M49FI1201.MV", "M49FI1201"),
        ("P-=AO2.3/M49DKA050.KEY:MAN", "M49DKA050.KEY", "M49DKA050"),
        ("=DI5.12/M49KN050.PWR:CALC_VAL", "M49KN050.PWR", "M49KN050"),
        ("=DO1.1/940M02M1.STRT:MAN", "940M02M1.STRT", "940M02M1"),
        ("=DO1.2/940M02M1.STRT:SELECTED", "940M02M1.STRT", "940M02M1"),
        ("=AI1.1/940LC391.MV:AUTO", "940LC391.MV", "940LC391"),
        ("=AO2.1/940FC400.OUT:REMOTE", "940FC400.OUT", "940FC400"),
        ("=DI3.1/940XV101.RUN:ENABLE", "940XV101.RUN", "940XV101"),
    ]
    for source, expected_device, expected_loop in cases:
        r = GrammarParser.parse_reference(source)
        assert r is not None, source
        assert r.device_tag == expected_device, source
        assert r.loop_tag == expected_loop, source
        assert ":" not in r.device_tag, source


def test_device_tag_preserves_full_engineering_suffixes():
    """Device Tags must be captured in full — no suffix whitelist, no length cap."""
    cases = [
        ("=DI1.1/940XA1899.FAULT", "940XA1899.FAULT", "940XA1899"),
        ("=DO2.1/945PB25M1.START2", "945PB25M1.START2", "945PB25M1"),
        ("=DI3.1/940XA1899.FAULT:ERR", "940XA1899.FAULT", "940XA1899"),
        ("=DO4.1/945PB25M1.START2:MAN", "945PB25M1.START2", "945PB25M1"),
        ("=AI1.1/940XA100.ALARM", "940XA100.ALARM", "940XA100"),
        ("=AO1.1/940XV200.CLOSE", "940XV200.CLOSE", "940XV200"),
        ("=DI1.1/M49M021.READY", "M49M021.READY", "M49M021"),
        ("=DO1.1/940M02M1.STOP", "940M02M1.STOP", "940M02M1"),
        ("=DI1.1/940M02M1.RUN", "940M02M1.RUN", "940M02M1"),
        ("=AI1.1/940LC391.MV", "940LC391.MV", "940LC391"),
        ("=AO1.1/940FC400.OUT", "940FC400.OUT", "940FC400"),
        ("=DO1.1/940XV101.SV1", "940XV101.SV1", "940XV101"),
        ("=DO1.2/940XV101.SV2", "940XV101.SV2", "940XV101"),
        ("=DO1.3/940XV101.GSO", "940XV101.GSO", "940XV101"),
        ("=DO1.4/940XV101.GSC", "940XV101.GSC", "940XV101"),
        ("=DI1.1/940XV101.OPEN", "940XV101.OPEN", "940XV101"),
        ("=DI1.2/940XV101.CLOSE", "940XV101.CLOSE", "940XV101"),
        ("=DI1.3/940XV101.SELECTED", "940XV101.SELECTED", "940XV101"),
        ("=DI1.4/940XV101.MAN", "940XV101.MAN", "940XV101"),
        ("=DI1.5/940XV101.AUTO", "940XV101.AUTO", "940XV101"),
        ("=DI1.6/940XV101.REMOTE", "940XV101.REMOTE", "940XV101"),
        ("=DI1.7/940XV101.TRIP", "940XV101.TRIP", "940XV101"),
        ("=DI1.8/940XV101.ALARM", "940XV101.ALARM", "940XV101"),
    ]
    for source, expected_device, expected_loop in cases:
        r = GrammarParser.parse_reference(source)
        assert r is not None, source
        assert r.device_tag == expected_device, f"{source} -> {r.device_tag}"
        assert r.loop_tag == expected_loop, source


def test_clean_device_tag_no_character_limit_or_suffix_list():
    """clean_device_tag must not truncate and must not rely on suffix whitelists."""
    assert GrammarParser.clean_device_tag("940XA1899.FAULT") == "940XA1899.FAULT"
    assert GrammarParser.clean_device_tag("945PB25M1.START2") == "945PB25M1.START2"
    assert GrammarParser.clean_device_tag("940XV101.SELECTED") == "940XV101.SELECTED"
    assert GrammarParser.clean_device_tag("940XV101.SELECTED:ERR") == "940XV101.SELECTED"
    # Exact as printed — no assumed short-suffix glue stripping
    assert GrammarParser.clean_device_tag("940LC391.MVBLOCK") == "940LC391.MVBLOCK"
    assert GrammarParser.clean_device_tag("945FC400.OUTXYZ") == "945FC400.OUTXYZ"
    # Very long suffix still preserved in full
    long_tag = "940XA1899." + ("A" * 40) + "12"
    assert GrammarParser.clean_device_tag(long_tag) == long_tag


def test_unsupported_categories_are_ignored():
    """AIC/AOC/DIC/DOC and other non-I/O families must not be parsed."""
    for ref in (
        "=AOC264:17/949DKA050.KEY:SELECTED",
        "=AIC793:55/M49KN050.PWR:CALC_VAL",
        "=DOC10:1/TAG.OUT",
        "=DIC5:2/TAG.IN",
        "=ACC1/TAG.X",
        "=AICT1.1/TAG.MV",
    ):
        assert GrammarParser.parse_reference(ref) is None


def test_dedup_collapses_colon_attribute_variants():
    """Device Tags that differ only by a colon attribute are the same engineering tag."""
    refs = [
        GrammarParser.parse_reference("=AI800_22.5/M49FI1201.MV"),
        GrammarParser.parse_reference("=AI800_22.5:22/M49FI1201.MV:ERR"),
        GrammarParser.parse_reference("=AO2.3/945FC400.OUT"),
    ]
    assert all(r is not None for r in refs)
    unique, dups = DuplicateDetector.deduplicate_references(refs)
    assert len(unique) == 2
    assert dups == 1
    tags = {r.device_tag for r in unique}
    assert tags == {"M49FI1201.MV", "945FC400.OUT"}
    assert all(":" not in t for t in tags)


def test_dedup_removes_exact_duplicates_only():
    refs = [
        GrammarParser.parse_reference("=AI1.1/940LC391.MV"),
        GrammarParser.parse_reference("=AI1.1/940LC391.MV"),
    ]
    unique, dups = DuplicateDetector.deduplicate_references(refs)
    assert len(unique) == 1
    assert dups == 1


def test_validator_rejects_unsupported_categories():
    obj = EngineeringIO(
        io_family="AOC",
        io_type="AOC",
        category="AOC",
        card_number=264,
        channel_number=0,
        loop_tag="949DKA050",
        device_tag="949DKA050.KEY",
        source_reference="=AOC264/949DKA050.KEY",
    )
    ok, errors = Validator.validate_object(obj)
    assert not ok
    assert any("category" in e.lower() or "family" in e.lower() for e in errors)


def test_validator_accepts_supported_zero_channel():
    obj = EngineeringIO(
        io_family="AO",
        io_type="AO",
        category="AO",
        card_number=199,
        channel_number=0,
        loop_tag="M49ARA104",
        device_tag="M49ARA104.CA41",
        source_reference="=AO199/M49ARA104.CA41",
    )
    ok, errors = Validator.validate_object(obj)
    assert ok, errors


def test_excel_generation_columns(tmp_path):
    obj1 = EngineeringIO(
        io_family="AI800_",
        io_type="AI",
        category="AI800",
        card_number=5,
        channel_number=10,
        loop_tag="82LIC660",
        device_tag="82LIC660.MV",
        description="Felt Water Tank",
    )
    obj2 = EngineeringIO(
        io_family="AO",
        io_type="AO",
        category="AO",
        card_number=2,
        channel_number=3,
        loop_tag="945FC400",
        device_tag="945FC400.OUT",
        description="",
    )

    out_file = str(tmp_path / "test_io_list.xlsx")
    ExcelGenerator.generate_excel([obj1, obj2], out_file)
    wb = openpyxl.load_workbook(out_file)
    assert wb.sheetnames == ["I_O_List", FUNCTION_BLOCK_SUMMARY_SHEET]
    ws = wb["I_O_List"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, 15)]
    assert headers == [
        "Sr. No.", "$(TAG)", "$(NAME_40)", "$(DEVICETAG)",
        "Slot/Card", "Channel",
        "AI", "AO", "DI", "DO", "AI800_", "AO800_", "DI800_", "DO800_",
    ]
    assert ws["A1"].fill.fgColor.rgb.endswith("1E293B")
    assert ws["A1"].font.name == "Calibri"
    assert ws["A1"].font.bold is True
    assert ws["A2"].font.name == "Calibri"
    assert ws["A2"].fill.fgColor.rgb.endswith("FFFFFF")
    assert ws["A3"].fill.fgColor.rgb.endswith("F8FAFC")

    # Row 1: AI800 → AI800_ = 1, others blank
    row2 = [ws.cell(row=2, column=c).value for c in range(1, 15)]
    assert row2[1] == "82LIC660"
    assert row2[3] == "82LIC660.MV"
    assert row2[4] == 5
    assert row2[5] == 10
    assert row2[6:14] == [None, None, None, None, 1, None, None, None]

    # Row 2: AO → AO = 1
    row3 = [ws.cell(row=3, column=c).value for c in range(1, 15)]
    assert row3[3] == "945FC400.OUT"
    assert row3[4] == 2
    assert row3[5] == 3
    assert row3[6:14] == [None, 1, None, None, None, None, None, None]

    summary = wb[FUNCTION_BLOCK_SUMMARY_SHEET]
    assert [summary.cell(row=1, column=c).value for c in range(1, 3)] == [
        "Functional Block",
        "Total Count",
    ]
    assert [summary.cell(row=r, column=1).value for r in range(2, 6)] == list(
        SUPPORTED_FUNCTION_BLOCKS
    )
    assert all(summary.cell(row=r, column=2).value == 0 for r in range(2, 6))


def test_category_mapper_indicators():
    from backend.pc_element.parser.category_mapper import (
        build_category_indicator_values,
        apply_category_columns,
        CATEGORY_INDICATOR_COLUMNS,
    )

    ai = build_category_indicator_values("AI")
    assert ai["AI"] == 1
    assert all(ai[c] == "" for c in CATEGORY_INDICATOR_COLUMNS if c != "AI")

    di800 = build_category_indicator_values("DI800_")
    assert di800["DI800_"] == 1
    assert di800["DI"] == ""

    row = apply_category_columns({
        "Device Tag": "940M02M1.RUN",
        "Category": "DI800_",
        "Slot/Card": 2,
    })
    assert "Category" not in row
    assert row["DI800_"] == 1
    assert row["Device Tag"] == "940M02M1.RUN"
    assert row["Slot/Card"] == 2
    assert [row[c] for c in CATEGORY_INDICATOR_COLUMNS] == ["", "", "", "", "", "", 1, ""]


def test_pc_element_parser_service_pipeline(tmp_path):
    from unittest.mock import patch
    from backend.pc_element.parser.parser_service import PCParserService
    from backend.pc_element.parser.pdf_reader import PageContent

    mock_pages = [
        PageContent(
            page_number=1,
            text=(
                "-AI800_2.1/M49M021.CURR\n"
                "=AO2.3/945FC400.OUT\n"
                "=DI5.12/940M03M1.RUN\n"
                "=AOC264:17/949DKA050.KEY:SELECTED\n"
                "=AIC793:55/M49KN050.PWR:CALC_VAL\n"
                "Felt Water Tank Level\n"
                "PM2\\Node22\n"
                "White Water System"
            ),
            raw_lines=[
                "-AI800_2.1/M49M021.CURR",
                "=AO2.3/945FC400.OUT",
                "=DI5.12/940M03M1.RUN",
                "=AOC264:17/949DKA050.KEY:SELECTED",
                "=AIC793:55/M49KN050.PWR:CALC_VAL",
                "Felt Water Tank Level",
                "PM2\\Node22",
                "White Water System",
            ]
        )
    ]

    with patch(
        "backend.pc_element.parser.parser_service.PDFReader.read_all_pages",
        return_value=mock_pages,
    ):
        service = PCParserService(
            file_path=__file__,
            job_id="test_job",
            output_dir=str(tmp_path),
        )
        res = service.execute_pipeline()

        assert len(res.errors) == 0
        assert res.total_io_found == 3  # AI800 + AO + DI; AOC/AIC skipped
        assert res.ai800_count == 1
        assert res.ao_count == 1
        assert res.di_count == 1
        assert res.aoc_count == 0
        assert res.aic_count == 0
        # Category column replaced by eight indicators; section order preserved
        assert "Category" not in res.preview_data[0]
        assert [row["AO"] for row in res.preview_data] == [1, "", ""]
        assert [row["DI"] for row in res.preview_data] == ["", 1, ""]
        assert [row["AI800_"] for row in res.preview_data] == ["", "", 1]
        assert set(res.preview_data[0].keys()) >= {
            "Sr. No.", "Loop Tag", "Description", "Device Tag",
            "AI", "AO", "DI", "DO", "AI800_", "AO800_", "DI800_", "DO800_",
            "Slot/Card", "Channel",
        }
        keys = list(res.preview_data[0].keys())
        device_idx = keys.index("Device Tag")
        assert keys[device_idx + 1] == "Slot/Card"
        assert keys[device_idx + 2] == "Channel"


@pytest.mark.integration
def test_reference_o2_pc32_full_coverage():
    """Every detectable I/O address in the user reference O2-PC32.pdf must be extracted."""
    import re
    import tempfile
    from backend.pc_element.parser.parser_service import PCParserService
    from backend.pc_element.parser.pdf_reader import PDFReader
    from backend.pc_element.parser.page_cleaner import PageCleaner
    from backend.pc_element.parser.io_reference_detector import IOReferenceDetector
    from backend.pc_element.parser.grammar_parser import GrammarParser
    from backend.pc_element.parser.duplicate_detector import DuplicateDetector
    from backend.pc_element.parser.validator import Validator, EngineeringIO

    pdf = r"c:\Users\vedan\Downloads\Testing Material\Project references\O2_LOGIC_PDF\O2-PC32.pdf"
    if not os.path.exists(pdf):
        pdf = r"backend\temp\uploads\0eb5589f-fbdf-4859-ad0b-edc9539aa4ec\O2-PC32.pdf"
    if not os.path.exists(pdf):
        pytest.skip("O2-PC32.pdf not available")

    # Include CARD.CHANNEL:TERMINAL/TAG form in ground truth — only 8 supported families
    STRICT = re.compile(
        r'''(?ix)
        (?:[-+]?\s*P\s*-?\s*=?\s*|[=+\-]+\s*)?
        (?P<prefix>AI800_|AO800_|DI800_|DO800_|AI800|AO800|DI800|DO800|AI|AO|DI|DO)
        \s*_?\s*(?P<card>\d{1,4})
        (?:
            \s*\.\s*(?P<channel>\d{1,3})\s*:\s*(?P<terminal>\d{1,4})
          | \s*\.\s*(?P<channel2>\d{1,3})
          | \s*:\s*(?P<port>\d{1,4})
        )?
        \s*/\s*
        (?P<tag>[A-Za-z0-9_][A-Za-z0-9_\-]*(?:\.[A-Za-z0-9_]+)?(?::[A-Za-z0-9_]+)?)
        (?![A-Za-z0-9_.])
        '''
    )

    pages = PDFReader(pdf).read_all_pages()
    gt = set()
    for page in pages:
        for src in [page.text] + page.raw_lines:
            for m in STRICT.finditer(src or ""):
                prefix = m.group("prefix").upper()
                if prefix in ("AI800", "AO800", "DI800", "DO800"):
                    prefix = prefix + "_"
                card = int(m.group("card"))
                ch = m.group("channel") or m.group("channel2") or m.group("port")
                channel = int(ch) if ch else 0
                tag = GrammarParser.clean_device_tag(m.group("tag"))
                gt.add((prefix, card, channel, tag))

    out = tempfile.mkdtemp()
    res = PCParserService(pdf, "cov_ref", out).execute_pipeline()
    assert not res.errors

    parsed = []
    for page in pages:
        cleaned = PageCleaner.clean_page_lines(page.raw_lines)
        for c in IOReferenceDetector.detect_candidates_in_page(page.text, cleaned):
            ref = GrammarParser.parse_reference(c, page.page_number)
            if ref:
                parsed.append(ref)
    unique, _ = DuplicateDetector.deduplicate_references(parsed)
    objs = [
        EngineeringIO(
            io_family=r.io_family, io_type=r.io_type, category=r.category,
            card_number=r.card_number, channel_number=r.channel_number,
            loop_tag=r.loop_tag, device_tag=r.device_tag,
            source_reference=r.source_reference, page_number=r.page_number,
        )
        for r in unique
    ]
    valid, _ = Validator.filter_and_validate_all(objs)
    out_keys = {
        (v.io_family.upper(), v.card_number, v.channel_number, v.device_tag.upper())
        for v in valid
    }

    missing = sorted(gt - out_keys)
    assert len(missing) == 0, f"Missing {len(missing)} refs, e.g. {missing[:15]}"
    assert res.total_io_found >= len(gt)
    # Device Tags must be clean engineering tags (no colon attributes)
    colon_tags = {k for k in out_keys if ":" in k[3]}
    assert len(colon_tags) == 0, f"Device Tags must not contain colon attributes: {list(colon_tags)[:10]}"


def test_function_block_declaration_counted():
    """Only PIDCON(...) style declarations increment the count."""
    pages = [
        "PIDCON(0,0,1,1,1,0)\nTRACKA\nPARAM1",
        "Some noise PIDCON (1,0,0)\nand MOTCON(0,1)",
    ]
    counts = count_function_blocks(pages)
    assert counts["PIDCON"] == 2
    assert counts["MOTCON"] == 1
    assert counts["VALVECON"] == 0
    assert counts["MANSTN"] == 0


def test_function_block_references_ignored():
    """Parameter / cross-reference labels must never count as declarations."""
    pages = [
        """
        =PIDCON1:94/940LC391:PARAM1
        =PIDCON1:32/940LC391:PARAM2
        PIDCON1:55/940LC391:PARAM4
        PIDCON1:86/940LC391:POUT
        PIDCON1:46/940LC391:PARAM3
        P-=MOTCON1:10/940M01:OUT
        VALVECON1:22/940XV101:SV1
        MANSTN1:5/940MS01:MAN
        DEFAULT PIDCON
        PIDCON1
        """
    ]
    counts = count_function_blocks(pages)
    assert counts == {
        "PIDCON": 0,
        "MOTCON": 0,
        "VALVECON": 0,
        "MANSTN": 0,
    }


def test_function_block_mixed_counts():
    """Mixed declarations produce the expected summary totals."""
    page = "\n".join(
        [
            "PIDCON(0,0,1,1,1,0)",
            "PIDCON(1,0,0,0,0,0)",
            "PIDCON(0,1,1,1,0,0)",
            "MOTCON(0,0)",
            "MOTCON(1,1)",
            "VALVECON(2,2)",
            "MANSTN(0)",
            "MANSTN(1)",
            "MANSTN(2)",
            "MANSTN(3)",
            "MANSTN(4)",
            # noise that must be ignored
            "=PIDCON1:94/940LC391:PARAM1",
            "MOTCON1:10/TAG:OUT",
        ]
    )
    counts = count_function_blocks([page])
    assert counts == {
        "PIDCON": 3,
        "MOTCON": 2,
        "VALVECON": 1,
        "MANSTN": 5,
    }
    rows = function_block_summary_rows(counts)
    assert rows == [
        {"Functional Block": "PIDCON", "Total Count": 3},
        {"Functional Block": "MOTCON", "Total Count": 2},
        {"Functional Block": "VALVECON", "Total Count": 1},
        {"Functional Block": "MANSTN", "Total Count": 5},
    ]


def test_function_block_summary_excel_sheet(tmp_path):
    """Excel export includes Function Block Summary with provided counts."""
    obj = EngineeringIO(
        io_family="AI",
        io_type="AI",
        category="AI",
        card_number=1,
        channel_number=1,
        loop_tag="940LC391",
        device_tag="940LC391.MV",
        description="Level",
    )
    counts = {"PIDCON": 3, "MOTCON": 2, "VALVECON": 1, "MANSTN": 5}
    out_file = str(tmp_path / "fb_summary.xlsx")
    ExcelGenerator.generate_excel([obj], out_file, function_block_counts=counts)
    wb = openpyxl.load_workbook(out_file)
    assert "I_O_List" in wb.sheetnames
    assert FUNCTION_BLOCK_SUMMARY_SHEET in wb.sheetnames

    summary = wb[FUNCTION_BLOCK_SUMMARY_SHEET]
    assert summary["A1"].value == "Functional Block"
    assert summary["B1"].value == "Total Count"
    assert summary["A1"].fill.fgColor.rgb.endswith("1E293B")
    expected = [
        ("PIDCON", 3),
        ("MOTCON", 2),
        ("VALVECON", 1),
        ("MANSTN", 5),
    ]
    for idx, (name, total) in enumerate(expected, start=2):
        assert summary.cell(row=idx, column=1).value == name
        assert summary.cell(row=idx, column=2).value == total
