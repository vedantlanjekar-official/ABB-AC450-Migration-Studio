"""Multi-file PC/DB jobs must download one workbook that contains every file."""

from __future__ import annotations

import shutil
import uuid
from pathlib import Path

import fitz
import openpyxl
import pytest

from backend.core.config import settings
from backend.services.conversion_service import ConversionService
from backend.services.job_manager import job_store

REPO = Path(__file__).resolve().parents[2]
SAMPLE_DB_PDF = REPO / "examples" / "sample_ac450_db.pdf"
SAMPLE_DB_BAX = Path(__file__).resolve().parent / "fixtures" / "sample_db.bax"
SAMPLE_AAX_DIR = REPO / ".AXX Data"


def _count_sheet_rows(excel_path: Path, sheet: str | None = None) -> int:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))


def _run_job(conversion_type: str, sources: list[tuple[Path, str]]) -> dict:
    job_id = f"test_multi_{conversion_type.lower()}_{uuid.uuid4().hex[:8]}"
    upload_dir = settings.UPLOAD_DIR / job_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    names = []
    for src, dest_name in sources:
        shutil.copy(src, upload_dir / dest_name)
        names.append(dest_name)
    job_store.create_job(job_id, names)
    ConversionService(job_id).run_conversion_pipeline(conversion_type)
    job = job_store.get_job(job_id)
    assert job["status"] == "completed", job.get("message")
    return job


def _write_pc_pdf(path: Path, tags: list[str]) -> None:
    doc = fitz.open()
    page = doc.new_page()
    body = "\n".join(tags + [
        "Felt Water Tank Level Control",
        "PM2 Node22 White Water System",
        "PIDCON (",
    ])
    page.insert_text((72, 72), body)
    doc.save(str(path))
    doc.close()


def test_db_multi_pdf_excel_contains_every_file():
    if not SAMPLE_DB_PDF.exists():
        pytest.skip("sample DB PDF missing")
    one = _run_job("DB", [(SAMPLE_DB_PDF, "only.pdf")])
    n_one = _count_sheet_rows(Path(one["excel_file_path"]))
    assert n_one >= 4
    job = _run_job(
        "DB",
        [(SAMPLE_DB_PDF, "part1.pdf"), (SAMPLE_DB_PDF, "part2.pdf")],
    )
    excel_path = Path(job["excel_file_path"])
    assert excel_path.name == "DB_Element.xlsx"
    n_all = _count_sheet_rows(excel_path)
    assert n_all == n_one * 2
    assert job["total_objects"] == n_all


def test_db_multi_bax_excel_contains_every_file():
    if not SAMPLE_DB_BAX.exists():
        pytest.skip("sample DB BAX missing")
    one = _run_job("DB", [(SAMPLE_DB_BAX, "only.bax")])
    n_one = _count_sheet_rows(Path(one["excel_file_path"]))
    assert n_one >= 4
    job = _run_job(
        "DB",
        [(SAMPLE_DB_BAX, "unit1.bax"), (SAMPLE_DB_BAX, "unit2.bax")],
    )
    excel_path = Path(job["excel_file_path"])
    assert excel_path.name == "DB_Element.xlsx"
    assert _count_sheet_rows(excel_path) == n_one * 2


def test_db_mixed_pdf_and_bax_excel_contains_every_file():
    if not SAMPLE_DB_PDF.exists() or not SAMPLE_DB_BAX.exists():
        pytest.skip("sample DB sources missing")
    pdf_job = _run_job("DB", [(SAMPLE_DB_PDF, "db.pdf")])
    bax_job = _run_job("DB", [(SAMPLE_DB_BAX, "db.bax")])
    n_pdf = _count_sheet_rows(Path(pdf_job["excel_file_path"]))
    n_bax = _count_sheet_rows(Path(bax_job["excel_file_path"]))
    job = _run_job(
        "DB",
        [(SAMPLE_DB_PDF, "db.pdf"), (SAMPLE_DB_BAX, "db.bax")],
    )
    excel_path = Path(job["excel_file_path"])
    assert excel_path.name == "DB_Element.xlsx"
    assert _count_sheet_rows(excel_path) == n_pdf + n_bax


def test_pc_multi_aax_excel_contains_every_file():
    f1 = SAMPLE_AAX_DIR / "23JA0401.AAX"
    f2 = SAMPLE_AAX_DIR / "23JA0501.AAX"
    if not f1.exists() or not f2.exists():
        pytest.skip("sample AAX missing")
    a = _run_job("PC", [(f1, f1.name)])
    b = _run_job("PC", [(f2, f2.name)])
    n1 = _count_sheet_rows(Path(a["excel_file_path"]), "I_O_List")
    n2 = _count_sheet_rows(Path(b["excel_file_path"]), "I_O_List")
    job = _run_job("PC", [(f1, f1.name), (f2, f2.name)])
    excel_path = Path(job["excel_file_path"])
    assert excel_path.name == "PC_Element_IO_List.xlsx"
    assert _count_sheet_rows(excel_path, "I_O_List") == n1 + n2
    assert "Function Block Summary" in openpyxl.load_workbook(excel_path).sheetnames


def test_pc_multi_pdf_excel_contains_every_file(tmp_path):
    pdf_a = tmp_path / "loop_a.pdf"
    pdf_b = tmp_path / "loop_b.pdf"
    _write_pc_pdf(pdf_a, ["=AI1.1/940LC391.MV"])
    _write_pc_pdf(pdf_b, ["=AO2.3/945FC400.OUT"])
    job = _run_job("PC", [(pdf_a, "loop_a.pdf"), (pdf_b, "loop_b.pdf")])
    excel_path = Path(job["excel_file_path"])
    assert excel_path.name == "PC_Element_IO_List.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    rows = list(wb["I_O_List"].iter_rows(min_row=2, values_only=True))
    tags = " ".join(str(c or "") for r in rows for c in r)
    assert "940LC391" in tags
    assert "945FC400" in tags
    assert len(rows) >= 2
