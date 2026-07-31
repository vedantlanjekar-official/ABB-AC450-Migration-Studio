import io
import time
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from backend.io_address_arrangement import IOAddressArranger
from backend.main import app
from backend.mapper.category_mapper import CATEGORY_INDICATOR_COLUMNS


client = TestClient(app)


def _build_generated_workbook(path: Path, record_count: int = 17) -> list[str]:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Clubbed_IO"
    headers = [
        "Tag",
        "Loop Tag",
        "NAME",
        "DESCR",
        *CATEGORY_INDICATOR_COLUMNS,
    ]
    worksheet.append(headers)

    tags = [f"949TI{i:03d}A.MV" for i in range(1, record_count + 1)]
    ai_col = headers.index("AI") + 1
    for index, tag in enumerate(tags, start=1):
        row = [f"AI1.{index}", tag.rsplit(".", 1)[0], tag, "Temperature", *("" for _ in CATEGORY_INDICATOR_COLUMNS)]
        worksheet.append(row)
        worksheet.cell(row=index + 1, column=ai_col, value=1)

    # A second category proves that every sequence starts independently.
    ao_row = ["AO1.1", "949TY001", "949TY001.OUT", "Output", *("" for _ in CATEGORY_INDICATOR_COLUMNS)]
    worksheet.append(ao_row)
    worksheet.cell(
        row=record_count + 2,
        column=headers.index("AO") + 1,
        value=1,
    )
    workbook.save(path)
    return tags


def test_arranger_preserves_tags_and_rolls_to_next_card_pair(tmp_path):
    source = tmp_path / "generated_db.xlsx"
    output = tmp_path / "arranged.xlsx"
    tags = _build_generated_workbook(source)

    result = IOAddressArranger().arrange(source, output)

    assert result.category_counts == {"AI": 17, "AO": 1}
    assert result.records_by_category["AI"][0].device_tag == tags[0]
    assert result.records_by_category["AI"][0].loop_tag == "949TI001A"

    workbook = openpyxl.load_workbook(output, data_only=True)
    try:
        assert workbook.sheetnames == ["AI", "AO"]
        ai_sheet = workbook["AI"]
        assert [ai_sheet.cell(1, col).value for col in range(1, 6)] == [
            "$(TAG)",
            "$(DEVICETAG)",
            None,
            "$(TAG)",
            "$(DEVICETAG)",
        ]
        assert ai_sheet["A1"].fill.fgColor.rgb.endswith("1E293B")
        assert ai_sheet["A1"].font.name == "Calibri"
        assert ai_sheet["A1"].font.bold is True
        assert ai_sheet["A2"].font.name == "Calibri"
        assert ai_sheet["A2"].fill.fgColor.rgb.endswith("FFFFFF")
        assert ai_sheet["A3"].fill.fgColor.rgb.endswith("F8FAFC")
        assert [ai_sheet.cell(2, col).value for col in range(1, 6)] == [
            "AI1.1",
            tags[0],
            None,
            "AI2.1",
            tags[0],
        ]
        assert ai_sheet["A17"].value == "AI1.16"
        assert ai_sheet["D17"].value == "AI2.16"
        assert ai_sheet["G2"].value == "AI3.1"
        assert ai_sheet["J2"].value == "AI4.1"
        assert ai_sheet["H2"].value == tags[16]
        assert ai_sheet["K2"].value == tags[16]
        assert ai_sheet["F2"].value is None
        assert ai_sheet["I2"].value is None

        ao_sheet = workbook["AO"]
        assert ao_sheet["A2"].value == "AO1.1"
        assert ao_sheet["D2"].value == "AO2.1"
        assert ao_sheet["B2"].value == "949TY001.OUT"
    finally:
        workbook.close()


def test_reader_streams_rows_without_random_cell_access(tmp_path, monkeypatch):
    """Large read-only workbooks must not degrade to quadratic cell lookups."""
    source = tmp_path / "generated_db.xlsx"
    _build_generated_workbook(source, record_count=64)

    def fail_random_access(*args, **kwargs):
        raise AssertionError("Read-only worksheet.cell() must not be used")

    monkeypatch.setattr(ReadOnlyWorksheet, "cell", fail_random_access)
    result = IOAddressArranger().read_records(source)

    assert result.source_records == 65
    assert result.category_counts == {"AI": 64, "AO": 1}


def test_arranger_reads_pc_device_tag_and_800_category(tmp_path):
    source = tmp_path / "generated_pc.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "I_O_List"
    worksheet.append(
        [
            "Sr. No.",
            "$(TAG)",
            "Description",
            "$(DEVICETAG)",
            *CATEGORY_INDICATOR_COLUMNS,
            "Slot/Card",
            "Channel",
        ]
    )
    worksheet.append(
        [1, "940M02M1", "", "940M02M1.RUN", "", "", "", "", "", "", 1, "", 2, 4]
    )
    workbook.save(source)

    result = IOAddressArranger().read_records(source)

    assert result.category_counts == {"DI800_": 1}
    record = result.records_by_category["DI800_"][0]
    assert record.device_tag == "940M02M1.RUN"
    assert record.loop_tag == "940M02M1"
    assert IOAddressArranger().generate_address_pair("DI800_", 0) == (
        "DI800_1.1",
        "DI800_2.1",
    )


@pytest.mark.parametrize(
    ("category", "last_first_pair_index", "next_pair_index", "last_channel", "next_pair"),
    [
        ("AI", 15, 16, 16, ("AI3.1", "AI4.1")),
        ("AO", 15, 16, 16, ("AO3.1", "AO4.1")),
        ("AI800_", 7, 8, 8, ("AI800_3.1", "AI800_4.1")),
        ("AO800_", 7, 8, 8, ("AO800_3.1", "AO800_4.1")),
        ("DI", 31, 32, 32, ("DI3.1", "DI4.1")),
        ("DO", 31, 32, 32, ("DO3.1", "DO4.1")),
        ("DI800_", 31, 32, 32, ("DI800_3.1", "DI800_4.1")),
        ("DO800_", 31, 32, 32, ("DO800_3.1", "DO800_4.1")),
    ],
)
def test_hardware_channel_limits_per_category(
    category,
    last_first_pair_index,
    next_pair_index,
    last_channel,
    next_pair,
):
    arranger = IOAddressArranger()

    assert arranger.generate_address_pair(category, last_first_pair_index) == (
        f"{category}1.{last_channel}",
        f"{category}2.{last_channel}",
    )
    assert arranger.generate_address_pair(category, next_pair_index) == next_pair


@pytest.mark.parametrize(
    ("category", "beyond_old_limit_index", "expected_left_card"),
    [
        ("AI", 256, 33),
        ("AO", 256, 33),
        ("AI800_", 128, 33),
        ("AO800_", 128, 33),
        ("DI", 512, 33),
        ("DO", 512, 33),
        ("DI800_", 512, 33),
        ("DO800_", 512, 33),
    ],
)
def test_address_generation_continues_past_card_32(
    category,
    beyond_old_limit_index,
    expected_left_card,
):
    arranger = IOAddressArranger()
    channels = {"AI": 16, "AO": 16, "AI800_": 8, "AO800_": 8, "DI": 32, "DO": 32, "DI800_": 32, "DO800_": 32}[
        category
    ]

    last_within_32 = arranger.generate_address_pair(category, beyond_old_limit_index - 1)
    assert last_within_32[0].startswith(f"{category}31.")
    assert last_within_32[1].startswith(f"{category}32.")

    next_pair = arranger.generate_address_pair(category, beyond_old_limit_index)
    assert next_pair == (
        f"{category}{expected_left_card}.1",
        f"{category}{expected_left_card + 1}.1",
    )
    assert beyond_old_limit_index % channels == 0


def test_arranger_accepts_dataset_beyond_former_32_card_capacity(tmp_path):
    source = tmp_path / "generated_db.xlsx"
    output = tmp_path / "arranged.xlsx"
    tags = _build_generated_workbook(source, record_count=257)

    result = IOAddressArranger().arrange(source, output)

    assert result.category_counts["AI"] == 257
    assert output.exists()

    workbook = openpyxl.load_workbook(output, data_only=True)
    try:
        ai_sheet = workbook["AI"]
        # 257 AI records => 17 card pairs => cards 33/34 for the overflow channel
        assert ai_sheet.cell(2, 97).value == "AI33.1"
        assert ai_sheet.cell(2, 98).value == tags[256]
        assert ai_sheet.cell(2, 100).value == "AI34.1"
        assert ai_sheet.cell(2, 101).value == tags[256]
    finally:
        workbook.close()


def test_io_arrangement_api_flow(tmp_path):
    source = tmp_path / "generated_db.xlsx"
    _build_generated_workbook(source, record_count=3)

    with source.open("rb") as file_handle:
        upload = client.post(
            "/api/upload",
            files={
                "files": (
                    source.name,
                    file_handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert upload.status_code == 200
    job_id = upload.json()["job_id"]

    process = client.post(
        "/api/process",
        json={"job_id": job_id, "conversion_type": "IO_ARRANGE"},
    )
    assert process.status_code == 200
    assert process.json()["conversion_type"] == "IO_ARRANGE"

    status = None
    for _ in range(30):
        response = client.get(f"/api/status/{job_id}")
        assert response.status_code == 200
        status = response.json()
        if status["status"] in {"completed", "failed"}:
            break
        time.sleep(0.1)

    assert status is not None
    assert status["status"] == "completed", status["errors"]
    assert status["conversion_type"] == "IO_ARRANGE"
    assert status["total_objects"] == 4
    assert status["generated_sheets"] == ["AI", "AO"]

    download = client.get(f"/api/download/{job_id}")
    assert download.status_code == 200
    assert "IO_Address_Arrangement.xlsx" in download.headers["content-disposition"]
    exported = openpyxl.load_workbook(io.BytesIO(download.content), data_only=True)
    try:
        assert exported["AI"]["A2"].value == "AI1.1"
        assert exported["AI"]["D2"].value == "AI2.1"
    finally:
        exported.close()
