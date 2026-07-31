"""Map generated DB/PC clubbed Excel rows into an ABB engineering import template."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

from backend.excel.design import build_db_excel_design
from backend.mapper.category_mapper import CATEGORY_INDICATOR_COLUMNS


SUPPORTED_CATEGORIES: Tuple[str, ...] = tuple(CATEGORY_INDICATOR_COLUMNS)
_CATEGORY_ALIASES = {
    category.rstrip("_"): category for category in SUPPORTED_CATEGORIES
}
_CATEGORY_ALIASES.update({category: category for category in SUPPORTED_CATEGORIES})

# Adjacent clubs are compatible when these unordered category pairs appear next to
# each other with the same Loop Tag. Slot assignment always puts the input first.
_COMPATIBLE_PAIRS = {
    frozenset({"AI", "AO"}): ("AI", "AO"),
    frozenset({"DO", "DI"}): ("DI", "DO"),
    frozenset({"AI800_", "AO800_"}): ("AI800_", "AO800_"),
    frozenset({"DO800_", "DI800_"}): ("DI800_", "DO800_"),
}

_DEVICE_TAG_HEADERS = ("$(DEVICETAG)", "DEVICE TAG", "NAME")
_LOOP_TAG_HEADERS = ("$(TAG)", "LOOP TAG")
_DESCRIPTION_HEADERS = ("$(NAME_40)", "DESCR", "DESCRIPTION")
_UNIT_HEADERS = ("$(DEVICETAG:UNIT)", "UNIT")
_MIN_HEADERS = ("$(DEVICETAG:MIN)", "RANGE MIN", "RANGEMIN")
_MAX_HEADERS = ("$(DEVICETAG:MAX)", "RANGE MAX", "RANGEMAX")

_OPTIONAL_SOURCE_HEADERS: Mapping[str, Tuple[str, ...]] = {
    "$(PACKAGE)": ("$(PACKAGE)", "PACKAGE"),
    "Process Area ID": ("PROCESS AREA ID", "PROCESS AREA", "AREA ID"),
    "$(EXE)": ("$(EXE)", "EXE"),
    "$(CTRLROOM)": ("$(CTRLROOM)", "CTRLROOM", "CONTROL ROOM"),
    "$(ALGROUP)": ("$(ALGROUP)", "ALGROUP", "ALARM GROUP"),
}

TEMPLATE_COLUMNS: Tuple[str, ...] = (
    "$(PACKAGE)",
    "Process Area ID",
    "$(EXE)",
    "$(CTRLROOM)",
    "$(ALGROUP)",
    "$(NAME40_1)",
    "$(TAG)",
    "$(CARDTYPE1)",
    "$(DEVICETAG1)",
    "$(DEVICETAG1:MIN)",
    "$(DEVICETAG1:MAX)",
    "$(DEVICETAG1:UNIT)",
    "$(CARDTYPE2)",
    "$(DEVICETAG2)",
    "$(DEVICETAG2:MIN)",
    "$(DEVICETAG2:MAX)",
    "$(DEVICETAG2:UNIT)",
)

_PREFERRED_SHEETS = ("Clubbed_IO", "I_O_List")
_MAX_HEADER_SCAN_ROWS = 120
_SHEET_TITLE = "Engineering_Template"


def _header_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split()).strip().upper()


def _canonical_category(value: Any) -> Optional[str]:
    text = _header_text(value).replace(" ", "")
    return _CATEGORY_ALIASES.get(text)


def _indicator_is_active(value: Any) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return _header_text(value) not in {"", "0", "FALSE", "NO", "N"}


def _blank(value: Any) -> Any:
    return "" if value is None else value


@dataclass(frozen=True)
class SourceRecord:
    category: str
    device_tag: Any
    loop_tag: Any = ""
    description: Any = ""
    unit: Any = ""
    range_min: Any = ""
    range_max: Any = ""
    optional_fields: Mapping[str, Any] = field(default_factory=dict)
    source_sheet: str = ""
    source_row: int = 0


@dataclass
class TemplateResult:
    template_rows: List[Dict[str, Any]] = field(default_factory=list)
    source_records: int = 0
    paired_clubs: int = 0
    singleton_rows: int = 0
    skipped_records: int = 0
    warnings: List[str] = field(default_factory=list)
    generated_sheets: List[str] = field(default_factory=list)
    preview_data: Dict[str, List[Dict[str, Any]]] = field(default_factory=dict)


@dataclass(frozen=True)
class _SheetSchema:
    header_row: int
    device_tag_col: int
    category_col: Optional[int]
    indicator_cols: Mapping[str, int]
    loop_tag_col: Optional[int]
    description_col: Optional[int]
    unit_col: Optional[int]
    min_col: Optional[int]
    max_col: Optional[int]
    optional_cols: Mapping[str, int]
    sheet_category: Optional[str]


class EngineeringTemplateGenerator:
    """Map already-clubbed DB/PC Excel rows into ABB engineering template rows."""

    def generate(
        self,
        input_path: Path | str,
        output_path: Path | str,
    ) -> TemplateResult:
        source = Path(input_path)
        destination = Path(output_path)
        if not source.exists():
            raise FileNotFoundError(f"Excel file not found: {source}")
        if source.suffix.lower() == ".xls":
            raise ValueError(
                "Legacy .xls files are not supported. Save the generated file as .xlsx."
            )

        records, skipped, warnings = self.read_records(source)
        if not records:
            raise ValueError(
                "No supported I/O records were found. Expected a Device Tag / NAME / "
                "$(DEVICETAG) column plus Category or AI/AO/DI/DO/800-series indicators."
            )

        template_rows, paired, singletons, map_warnings = self.map_records(records)
        warnings.extend(map_warnings)

        destination.parent.mkdir(parents=True, exist_ok=True)
        self._write_workbook(template_rows, destination)

        return TemplateResult(
            template_rows=template_rows,
            source_records=len(records),
            paired_clubs=paired,
            singleton_rows=singletons,
            skipped_records=skipped,
            warnings=warnings,
            generated_sheets=[_SHEET_TITLE],
            preview_data={_SHEET_TITLE: template_rows[:100]},
        )

    def read_records(
        self,
        input_path: Path | str,
    ) -> Tuple[List[SourceRecord], int, List[str]]:
        """Read clubbed workbook rows in sheet order without regrouping."""
        source = Path(input_path)
        workbook = openpyxl.load_workbook(source, data_only=False, read_only=True)
        records: List[SourceRecord] = []
        skipped = 0
        warnings: List[str] = []

        try:
            worksheets = self._preferred_worksheets(workbook)
            for worksheet in worksheets:
                schema = self._find_sheet_schema(worksheet)
                if schema is None:
                    continue

                for row_number, values in enumerate(
                    worksheet.iter_rows(
                        min_row=schema.header_row + 1,
                        values_only=True,
                    ),
                    start=schema.header_row + 1,
                ):
                    device_tag = self._row_value(values, schema.device_tag_col)
                    if device_tag is None or (
                        isinstance(device_tag, str) and not device_tag.strip()
                    ):
                        continue

                    category = self._category_for_values(values, schema)
                    if category is None:
                        skipped += 1
                        continue

                    optional_fields = {
                        key: self._row_value(values, col)
                        for key, col in schema.optional_cols.items()
                    }
                    records.append(
                        SourceRecord(
                            category=category,
                            device_tag=device_tag,
                            loop_tag=_blank(
                                self._row_value(values, schema.loop_tag_col)
                                if schema.loop_tag_col
                                else ""
                            ),
                            description=_blank(
                                self._row_value(values, schema.description_col)
                                if schema.description_col
                                else ""
                            ),
                            unit=_blank(
                                self._row_value(values, schema.unit_col)
                                if schema.unit_col
                                else ""
                            ),
                            range_min=_blank(
                                self._row_value(values, schema.min_col)
                                if schema.min_col
                                else ""
                            ),
                            range_max=_blank(
                                self._row_value(values, schema.max_col)
                                if schema.max_col
                                else ""
                            ),
                            optional_fields=optional_fields,
                            source_sheet=worksheet.title,
                            source_row=row_number,
                        )
                    )
                # Prefer the first sheet that actually yields engineering rows.
                if records:
                    break
            else:
                warnings.append(
                    "No Clubbed_IO / I_O_List sheet with supported engineering "
                    "headers was found."
                )
        finally:
            workbook.close()

        return records, skipped, warnings

    def map_records(
        self,
        records: Sequence[SourceRecord],
    ) -> Tuple[List[Dict[str, Any]], int, int, List[str]]:
        """Collapse adjacent compatible clubs into one template row each."""
        template_rows: List[Dict[str, Any]] = []
        paired = 0
        singletons = 0
        warnings: List[str] = []
        index = 0

        while index < len(records):
            current = records[index]
            nxt = records[index + 1] if index + 1 < len(records) else None
            slots = self._pair_slots(current, nxt)

            if slots is not None:
                primary, secondary = slots
                template_rows.append(self._build_template_row(primary, secondary))
                paired += 1
                index += 2
                continue

            if (
                nxt is not None
                and frozenset({current.category, nxt.category}) in _COMPATIBLE_PAIRS
                and self._normalize_loop(current.loop_tag)
                != self._normalize_loop(nxt.loop_tag)
            ):
                warnings.append(
                    f"Adjacent {current.category}/{nxt.category} rows at source "
                    f"rows {current.source_row}/{nxt.source_row} have different "
                    "Loop Tags; kept as separate template rows."
                )

            template_rows.append(self._build_template_row(current, None))
            singletons += 1
            index += 1

        return template_rows, paired, singletons, warnings

    @staticmethod
    def _normalize_loop(value: Any) -> str:
        return str(value or "").strip().upper()

    @classmethod
    def _pair_slots(
        cls,
        first: SourceRecord,
        second: Optional[SourceRecord],
    ) -> Optional[Tuple[SourceRecord, SourceRecord]]:
        if second is None:
            return None
        ordered = _COMPATIBLE_PAIRS.get(frozenset({first.category, second.category}))
        if ordered is None:
            return None
        if cls._normalize_loop(first.loop_tag) != cls._normalize_loop(second.loop_tag):
            return None
        if not cls._normalize_loop(first.loop_tag):
            return None

        by_category = {first.category: first, second.category: second}
        primary_category, secondary_category = ordered
        return by_category[primary_category], by_category[secondary_category]

    @staticmethod
    def _build_template_row(
        primary: SourceRecord,
        secondary: Optional[SourceRecord],
    ) -> Dict[str, Any]:
        row: Dict[str, Any] = {column: "" for column in TEMPLATE_COLUMNS}

        for key, value in primary.optional_fields.items():
            if key in row and value not in (None, ""):
                row[key] = value

        row["$(NAME40_1)"] = _blank(primary.description)
        row["$(TAG)"] = _blank(primary.loop_tag)
        row["$(CARDTYPE1)"] = primary.category
        row["$(DEVICETAG1)"] = _blank(primary.device_tag)
        row["$(DEVICETAG1:MIN)"] = _blank(primary.range_min)
        row["$(DEVICETAG1:MAX)"] = _blank(primary.range_max)
        row["$(DEVICETAG1:UNIT)"] = _blank(primary.unit)

        if secondary is not None:
            # Prefer secondary optional fields only when primary left them blank.
            for key, value in secondary.optional_fields.items():
                if key in row and row[key] in ("", None) and value not in (None, ""):
                    row[key] = value
            if not row["$(NAME40_1)"] and secondary.description not in (None, ""):
                row["$(NAME40_1)"] = secondary.description
            row["$(CARDTYPE2)"] = secondary.category
            row["$(DEVICETAG2)"] = _blank(secondary.device_tag)
            row["$(DEVICETAG2:MIN)"] = _blank(secondary.range_min)
            row["$(DEVICETAG2:MAX)"] = _blank(secondary.range_max)
            row["$(DEVICETAG2:UNIT)"] = _blank(secondary.unit)

        return row

    @staticmethod
    def _preferred_worksheets(workbook) -> List[Any]:
        by_title = {ws.title: ws for ws in workbook.worksheets}
        ordered: List[Any] = []
        for name in _PREFERRED_SHEETS:
            if name in by_title:
                ordered.append(by_title[name])
        for worksheet in workbook.worksheets:
            if worksheet not in ordered:
                ordered.append(worksheet)
        return ordered

    def _find_sheet_schema(self, worksheet) -> Optional[_SheetSchema]:
        max_row = min(worksheet.max_row or 1, _MAX_HEADER_SCAN_ROWS)
        sheet_category = _canonical_category(worksheet.title)
        candidates: List[Tuple[int, int, _SheetSchema]] = []

        for row_number, values in enumerate(
            worksheet.iter_rows(max_row=max_row, values_only=True),
            start=1,
        ):
            headers: Dict[str, int] = {}
            for col_number, value in enumerate(values, start=1):
                text = _header_text(value)
                if text and text not in headers:
                    headers[text] = col_number

            device_col = next(
                (headers[name] for name in _DEVICE_TAG_HEADERS if name in headers),
                None,
            )
            if device_col is None:
                continue

            category_col = headers.get("CATEGORY")
            indicator_cols = {
                category: headers[category]
                for category in SUPPORTED_CATEGORIES
                if category in headers
            }
            if category_col is None and not indicator_cols and sheet_category is None:
                continue

            optional_cols = {}
            for template_key, aliases in _OPTIONAL_SOURCE_HEADERS.items():
                col = next((headers[name] for name in aliases if name in headers), None)
                if col is not None:
                    optional_cols[template_key] = col

            schema = _SheetSchema(
                header_row=row_number,
                device_tag_col=device_col,
                category_col=category_col,
                indicator_cols=indicator_cols,
                loop_tag_col=next(
                    (headers[name] for name in _LOOP_TAG_HEADERS if name in headers),
                    None,
                ),
                description_col=next(
                    (
                        headers[name]
                        for name in _DESCRIPTION_HEADERS
                        if name in headers
                    ),
                    None,
                ),
                unit_col=next(
                    (headers[name] for name in _UNIT_HEADERS if name in headers),
                    None,
                ),
                min_col=next(
                    (headers[name] for name in _MIN_HEADERS if name in headers),
                    None,
                ),
                max_col=next(
                    (headers[name] for name in _MAX_HEADERS if name in headers),
                    None,
                ),
                optional_cols=optional_cols,
                sheet_category=sheet_category,
            )
            score = (
                (3 if category_col is not None else 0)
                + len(indicator_cols)
                + (2 if sheet_category else 0)
                + (1 if schema.loop_tag_col else 0)
            )
            candidates.append((-score, row_number, schema))

        if not candidates:
            return None
        candidates.sort(key=lambda item: (item[0], item[1]))
        return candidates[0][2]

    @staticmethod
    def _row_value(values: Tuple[Any, ...], col_number: int) -> Any:
        index = col_number - 1
        return values[index] if 0 <= index < len(values) else None

    @classmethod
    def _category_for_values(
        cls,
        values: Tuple[Any, ...],
        schema: _SheetSchema,
    ) -> Optional[str]:
        if schema.category_col:
            category = _canonical_category(
                cls._row_value(values, schema.category_col)
            )
            if category:
                return category

        for category in SUPPORTED_CATEGORIES:
            col_number = schema.indicator_cols.get(category)
            if col_number and _indicator_is_active(
                cls._row_value(values, col_number)
            ):
                return category

        return schema.sheet_category

    def _write_workbook(
        self,
        rows: Sequence[Mapping[str, Any]],
        output_path: Path,
    ) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = _SHEET_TITLE
        worksheet.freeze_panes = "A2"
        design = build_db_excel_design()

        for col_number, header in enumerate(TEMPLATE_COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=col_number, value=header)
            cell.fill = design.header_fill
            cell.font = design.header_font
            cell.alignment = design.center_align
            cell.border = design.thin_border
        worksheet.row_dimensions[1].height = 26

        for row_index, row in enumerate(rows, start=2):
            fill = design.zebra_fill if row_index % 2 == 0 else design.white_fill
            for col_number, header in enumerate(TEMPLATE_COLUMNS, start=1):
                value = row.get(header, "")
                if isinstance(value, str):
                    clean = value.lstrip("=+@").strip()
                    if clean.startswith("-") and len(clean) > 1 and clean[1].isalpha():
                        clean = clean[1:].strip()
                    value = clean
                    cell = worksheet.cell(row=row_index, column=col_number, value=value)
                    cell.data_type = "s"
                else:
                    cell = worksheet.cell(
                        row=row_index,
                        column=col_number,
                        value="" if value is None else value,
                    )
                cell.fill = fill
                cell.font = design.cell_font
                cell.border = design.thin_border
                text = str(cell.value or "")
                if len(text) < 15 and " " not in text:
                    cell.alignment = design.center_align
                else:
                    cell.alignment = design.left_align

        for col_number, header in enumerate(TEMPLATE_COLUMNS, start=1):
            max_len = len(header)
            for row_number in range(2, len(rows) + 2):
                cell_val = str(
                    worksheet.cell(row=row_number, column=col_number).value or ""
                )
                if len(cell_val) > max_len:
                    max_len = len(cell_val)
            worksheet.column_dimensions[get_column_letter(col_number)].width = min(
                max(max_len + 4, 12),
                60,
            )

        if rows:
            last_col = get_column_letter(len(TEMPLATE_COLUMNS))
            worksheet.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

        workbook.save(output_path)
