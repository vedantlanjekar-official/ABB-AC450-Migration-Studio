import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List
from backend.models.pc_element import PCElement
from backend.core.logging import get_logger

class PCExcelGenerator:
    """
    PC Element OpenPyXL Excel Generator Module.
    Generates a single-sheet Valmet Excel reference file sorted by Category -> Card -> Channel.
    Columns: Sr No, Loop Tag, Description, Device Tag, Category, Card Number, Channel Number.
    """

    COLUMNS = [
        "Sr No", "Loop Tag", "Description", "Device Tag",
        "Category", "Card Number", "Channel Number"
    ]

    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    DATA_FONT = Font(name="Calibri", size=10, color="000000")
    ZEBRA_FILL = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    def __init__(self, job_id: str = None):
        self.logger = get_logger(job_id)

    def generate_excel(self, elements: List[PCElement], output_path: Path) -> Path:
        """
        Sorts elements by Category -> Card -> Channel and writes openpyxl Excel file.
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Sort elements by Category, numeric Card Number, and numeric Channel Number
        sorted_elements = sorted(
            elements,
            key=lambda e: (
                e.category.upper(),
                self._safe_int(e.card_number),
                self._safe_int(e.channel_number),
                e.device_tag.upper()
            )
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Valmet PC Export"
        ws.views.sheetView[0].showGridLines = True

        # Write Header Row
        for col_idx, col_name in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER

        ws.row_dimensions[1].height = 26

        # Write Data Rows
        for row_idx, elem in enumerate(sorted_elements, start=2):
            elem.sr_no = row_idx - 1  # Auto-re-number after sorting
            row_dict = elem.to_dict()

            is_even = (row_idx % 2 == 0)

            for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                val = row_dict.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = self.DATA_FONT
                cell.border = self.THIN_BORDER

                if is_even:
                    cell.fill = self.ZEBRA_FILL

                # Alignment rules
                if col_name in ("Sr No", "Card Number", "Channel Number"):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name in ("Loop Tag", "Device Tag", "Category"):
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            ws.row_dimensions[row_idx].height = 20

        # Auto-fit Column Widths
        for col_idx in range(1, len(self.COLUMNS) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                cell_val = str(ws.cell(row=row, column=col_idx).value or "")
                if len(cell_val) > max_len:
                    max_len = len(cell_val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        wb.save(output_path)
        self.logger.info(f"PCExcelGenerator saved {len(sorted_elements)} sorted record(s) to {output_path.name}.")
        return output_path

    def _safe_int(self, val_str: str) -> int:
        try:
            return int(re.sub(r'\D', '', val_str))
        except Exception:
            return 9999
