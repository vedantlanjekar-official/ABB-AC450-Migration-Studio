"""Read generated DB/PC workbooks and create ABB address-pair worksheets."""

from __future__ import annotations

from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

from backend.excel.design import build_db_excel_design
from backend.mapper.category_mapper import CATEGORY_INDICATOR_COLUMNS


SUPPORTED_CATEGORIES: Tuple[str, ...] = tuple(CATEGORY_INDICATOR_COLUMNS)
CHANNELS_PER_CARD: Mapping[str, int] = {
    "AI": 16,
    "AO": 16,
    "AI800_": 8,
    "AO800_": 8,
    "DI": 32,
    "DO": 32,
    "DI800_": 32,
    "DO800_": 32,
}
# Card count grows with dataset size (odd/even duplicate pairs). Excel's
# column limit is the practical ceiling, not a fixed ABB card cap.
_EXCEL_MAX_COLUMNS = 16384
_CATEGORY_ALIASES = {
    category.rstrip("_"): category for category in SUPPORTED_CATEGORIES
}
_CATEGORY_ALIASES.update({category: category for category in SUPPORTED_CATEGORIES})
_DEVICE_TAG_HEADERS = ("DEVICE TAG", "NAME", "$(DEVICETAG)")
_LOOP_TAG_HEADERS = ("LOOP TAG", "$(TAG)")
_MAX_HEADER_SCAN_ROWS = 120


def _header_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split()).strip().upper()


def _canonical_category(value: Any) -> Optional[str]:
    text = _header_text(value).replace(" ", "")
    return _CATEGORY_ALIASES.get(text)


def _indicator_is_active(value: Any) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return _header_text(value) not in {"", "0", "FALSE", "NO", "N"}


@dataclass(frozen=True)
class IORecord:
    category: str
    device_tag: Any
    loop_tag: Any = ""
    source_sheet: str = ""
    source_row: int = 0


@dataclass
class ArrangementResult:
    records_by_category: Dict[str, List[IORecord]] = field(default_factory=dict)
    source_records: int = 0
    skipped_records: int = 0
    generated_sheets: List[str] = field(default_factory=list)
    preview_data: Dict[str, List[Dict[str, Any]]] = field(default_factory=dict)

    @property
    def category_counts(self) -> Dict[str, int]:
        return {
            category: len(records)
            for category, records in self.records_by_category.items()
        }


@dataclass(frozen=True)
class _SheetSchema:
    header_row: int
    device_tag_col: int
    category_col: Optional[int]
    indicator_cols: Mapping[str, int]
    loop_tag_col: Optional[int]
    sheet_category: Optional[str]


class IOAddressArranger:
    """Arrange existing generated engineering records without changing their tags."""

    def arrange(
        self,
        input_path: Path | str,
        output_path: Path | str,
    ) -> ArrangementResult:
        source = Path(input_path)
        destination = Path(output_path)
        if not source.exists():
            raise FileNotFoundError(f"Excel file not found: {source}")
        if source.suffix.lower() == ".xls":
            raise ValueError(
                "Legacy .xls files are not supported. Save the generated file as .xlsx."
            )

        result = self.read_records(source)
        if result.source_records == 0:
            raise ValueError(
                "No supported I/O records were found. Expected a Device Tag or NAME "
                "column plus Category/AI/AO/DI/DO/800-series category indicators."
            )

        destination.parent.mkdir(parents=True, exist_ok=True)
        self._write_workbook(result, destination)
        return result

    def read_records(self, input_path: Path | str) -> ArrangementResult:
        """Read all supported records while retaining source order per category."""
        source = Path(input_path)
        workbook = openpyxl.load_workbook(source, data_only=False, read_only=True)
        grouped: Dict[str, List[IORecord]] = OrderedDict(
            (category, []) for category in SUPPORTED_CATEGORIES
        )
        skipped = 0

        try:
            for worksheet in workbook.worksheets:
                schema = self._find_sheet_schema(worksheet)
                if schema is None:
                    continue

                for row_number, values in enumerate(
                    worksheet.iter_rows(
                        min_row=schema.header_row + 1,
                        values_only=True,
                    ),
                    start=schema.header_row + 1,
                ):
                    device_tag = self._row_value(values, schema.device_tag_col)
                    if device_tag is None or (
                        isinstance(device_tag, str) and not device_tag.strip()
                    ):
                        continue

                    category = self._category_for_values(values, schema)
                    if category is None:
                        skipped += 1
                        continue

                    loop_tag = (
                        self._row_value(values, schema.loop_tag_col)
                        if schema.loop_tag_col
                        else ""
                    )
                    grouped[category].append(
                        IORecord(
                            category=category,
                            device_tag=device_tag,
                            loop_tag=loop_tag if loop_tag is not None else "",
                            source_sheet=worksheet.title,
                            source_row=row_number,
                        )
                    )
        finally:
            workbook.close()

        populated = OrderedDict(
            (category, records)
            for category, records in grouped.items()
            if records
        )
        return ArrangementResult(
            records_by_category=dict(populated),
            source_records=sum(len(records) for records in populated.values()),
            skipped_records=skipped,
        )

    def generate_address_pair(self, category: str, index: int) -> Tuple[str, str]:
        """Return paired ABB card addresses for a zero-based record index."""
        canonical = _canonical_category(category)
        if canonical is None:
            raise ValueError(f"Unsupported engineering category: {category}")
        if index < 0:
            raise ValueError("Record index cannot be negative.")

        channels_per_card = CHANNELS_PER_CARD[canonical]
        card_pair_index, channel_index = divmod(index, channels_per_card)
        left_card = card_pair_index * 2 + 1
        right_card = left_card + 1
        channel = channel_index + 1
        return (
            f"{canonical}{left_card}.{channel}",
            f"{canonical}{right_card}.{channel}",
        )

    def _find_sheet_schema(self, worksheet) -> Optional[_SheetSchema]:
        max_row = min(worksheet.max_row or 1, _MAX_HEADER_SCAN_ROWS)
        sheet_category = _canonical_category(worksheet.title)
        candidates: List[Tuple[int, int, _SheetSchema]] = []

        for row_number, values in enumerate(
            worksheet.iter_rows(max_row=max_row, values_only=True),
            start=1,
        ):
            headers: Dict[str, int] = {}
            for col_number, value in enumerate(values, start=1):
                text = _header_text(value)
                if text and text not in headers:
                    headers[text] = col_number

            device_col = next(
                (headers[name] for name in _DEVICE_TAG_HEADERS if name in headers),
                None,
            )
            if device_col is None:
                continue

            category_col = headers.get("CATEGORY")
            indicator_cols = {
                category: headers[category]
                for category in SUPPORTED_CATEGORIES
                if category in headers
            }
            if category_col is None and not indicator_cols and sheet_category is None:
                continue

            schema = _SheetSchema(
                header_row=row_number,
                device_tag_col=device_col,
                category_col=category_col,
                indicator_cols=indicator_cols,
                loop_tag_col=next(
                    (
                        headers[name]
                        for name in _LOOP_TAG_HEADERS
                        if name in headers
                    ),
                    None,
                ),
                sheet_category=sheet_category,
            )
            score = (
                (3 if category_col is not None else 0)
                + len(indicator_cols)
                + (2 if sheet_category else 0)
            )
            candidates.append((-score, row_number, schema))

        if not candidates:
            return None
        candidates.sort(key=lambda item: (item[0], item[1]))
        return candidates[0][2]

    @staticmethod
    def _row_value(values: Tuple[Any, ...], col_number: int) -> Any:
        """Read a one-based column safely from a streamed worksheet row."""
        index = col_number - 1
        return values[index] if 0 <= index < len(values) else None

    @classmethod
    def _category_for_values(
        cls,
        values: Tuple[Any, ...],
        schema: _SheetSchema,
    ) -> Optional[str]:
        if schema.category_col:
            category = _canonical_category(
                cls._row_value(values, schema.category_col)
            )
            if category:
                return category

        for category in SUPPORTED_CATEGORIES:
            col_number = schema.indicator_cols.get(category)
            if col_number and _indicator_is_active(
                cls._row_value(values, col_number)
            ):
                return category

        return schema.sheet_category

    def _write_workbook(
        self,
        result: ArrangementResult,
        output_path: Path,
    ) -> None:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        design = build_db_excel_design()

        for category, records in result.records_by_category.items():
            worksheet = workbook.create_sheet(category)
            worksheet.freeze_panes = "A2"
            channels_per_card = CHANNELS_PER_CARD[category]
            card_pair_count = (
                len(records) + channels_per_card - 1
            ) // channels_per_card
            total_cards = card_pair_count * 2
            last_column = total_cards * 3 - 1
            if last_column > _EXCEL_MAX_COLUMNS:
                raise ValueError(
                    f"{category}: {len(records)} records require {total_cards} "
                    f"cards ({last_column} columns), which exceeds Excel's "
                    f"{_EXCEL_MAX_COLUMNS}-column limit."
                )
            worksheet.auto_filter.ref = (
                f"A1:{get_column_letter(last_column)}"
                f"{channels_per_card + 1}"
            )
            preview_rows: List[Dict[str, Any]] = []
            for card_index in range(total_cards):
                address_column = card_index * 3 + 1
                tag_column = address_column + 1
                separator_column = address_column + 2
                pair_index = card_index // 2
                source_start = pair_index * channels_per_card
                card_number = card_index + 1
                for col_number, header in (
                    (address_column, "$(TAG)"),
                    (tag_column, "$(DEVICETAG)"),
                    (separator_column, ""),
                ):
                    cell = worksheet.cell(row=1, column=col_number, value=header)
                    cell.fill = (
                        design.zebra_fill
                        if col_number == separator_column
                        else design.header_fill
                    )
                    cell.font = design.header_font
                    cell.alignment = design.center_align
                    cell.border = design.thin_border

                for channel_index in range(channels_per_card):
                    source_index = source_start + channel_index
                    if source_index >= len(records):
                        break
                    record = records[source_index]
                    address = f"{category}{card_number}.{channel_index + 1}"
                    excel_row = channel_index + 2
                    for col_number, value in (
                        (address_column, address),
                        (tag_column, record.device_tag),
                    ):
                        cell = worksheet.cell(
                            row=excel_row, column=col_number, value=value
                        )
                        cell.font = design.cell_font
                        cell.border = design.thin_border
                        cell.alignment = (
                            design.center_align
                            if col_number == address_column
                            else design.left_align
                        )
                        cell.fill = (
                            design.zebra_fill
                            if channel_index % 2 == 1
                            else design.white_fill
                        )

                    separator_cell = worksheet.cell(
                        row=excel_row, column=separator_column, value=""
                    )
                    separator_cell.font = design.cell_font
                    separator_cell.border = design.thin_border
                    separator_cell.fill = design.zebra_fill
                    separator_cell.alignment = design.center_align

                    while len(preview_rows) <= channel_index:
                        preview_rows.append({})
                    preview_rows[channel_index][f"$(TAG) {card_number}"] = address
                    preview_rows[channel_index][
                        f"$(DEVICETAG) {card_number}"
                    ] = record.device_tag

                worksheet.column_dimensions[
                    get_column_letter(address_column)
                ].width = 18
                worksheet.column_dimensions[get_column_letter(tag_column)].width = 32
                worksheet.column_dimensions[
                    get_column_letter(separator_column)
                ].width = 4

            for row_number in range(2, channels_per_card + 2):
                for col_number in range(1, last_column + 1):
                    cell = worksheet.cell(row=row_number, column=col_number)
                    if cell.value is None:
                        cell.border = design.thin_border
                        cell.font = design.cell_font
                        cell.alignment = design.center_align
                        if col_number % 3 == 0:
                            cell.fill = design.zebra_fill
                        else:
                            cell.fill = (
                                design.zebra_fill
                                if row_number % 2 == 1
                                else design.white_fill
                            )

            worksheet.row_dimensions[1].height = 26
            result.generated_sheets.append(category)
            result.preview_data[category] = preview_rows[:100]

        workbook.save(output_path)
