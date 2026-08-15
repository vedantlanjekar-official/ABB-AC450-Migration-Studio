"""Compare AaxReader synth vs Excel rows for one or more AAX files."""
from __future__ import annotations

import re
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from backend.pc_element.parser.aax_reader import AaxReader  # noqa: E402
from backend.pc_element.parser.io_reference_detector import IOReferenceDetector  # noqa: E402
from backend.pc_element.parser.grammar_parser import GrammarParser  # noqa: E402
from backend.pc_element.parser.parser_service import PCParserService  # noqa: E402
import openpyxl  # noqa: E402


def diagnose(path: Path) -> None:
    pages = AaxReader(str(path)).read_all_pages()
    synth = []
    for p in pages:
        synth.extend((p.text_layers or {}).get("aax_synth", "").splitlines())
    synth = [s for s in synth if s.strip()]

    parsed = []
    skipped = []
    for p in pages:
        cands = IOReferenceDetector.detect_candidates_in_page(
            p.text, p.raw_lines, text_layers=p.text_layers
        )
        for c in cands:
            refs = GrammarParser.parse_all_references(c, page_number=p.page_number)
            if refs:
                parsed.extend(refs)
            else:
                skipped.append(c)

    with tempfile.TemporaryDirectory() as tmp:
        res = PCParserService(str(path), "diag", tmp).execute_pipeline()
        wb = openpyxl.load_workbook(res.excel_file_path)
        ws = wb["I_O_List"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        desc_idx = headers.index("Description") if "Description" in headers else 2
        tag_idx = headers.index("Device Tag") if "Device Tag" in headers else 3
        blank_desc = sum(1 for r in rows if not (r[desc_idx] or "").strip())
        tags = {(r[tag_idx] or "").upper() for r in rows if r[tag_idx]}

    print(f"=== {path.name} ===")
    print(f"  synth candidates     : {len(synth)}")
    print(f"  grammar parsed       : {len(parsed)}  skipped={len(skipped)}")
    print(f"  pipeline records     : {res.successfully_parsed}")
    print(f"  excel rows           : {len(rows)}")
    print(f"  excel blank desc     : {blank_desc}/{len(rows)}")
    print(f"  unique excel tags    : {len(tags)}")
    print(f"  invalid skipped      : {res.invalid_references}")
    print(f"  duplicates removed   : {res.duplicates_removed}")
    if skipped[:8]:
        print("  grammar skip sample  :")
        for s in skipped[:8]:
            print(f"    {s[:100]}")
    synth_tags = []
    for s in synth:
        if "/" in s:
            synth_tags.append(s.split("/", 1)[1].upper())
    missing = sorted(set(synth_tags) - tags)
    print(f"  synth tags missing from excel: {len(missing)}")
    for t in missing[:20]:
        print(f"    {t}")
    print()


def main() -> int:
    files = sys.argv[1:]
    root = ROOT / ".AXX Data"
    if not files:
        files = [
            str(root / "23JA1601.AAX"),
            str(root / "23JA0501.AAX"),
            str(root / "23JA1801.AAX"),
            str(root / "23JA0401.AAX"),
        ]
    for f in files:
        diagnose(Path(f))
    return 0


if __name__ == "__main__":
    sys.exit(main())
